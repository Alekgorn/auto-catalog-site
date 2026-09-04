import base64
import gzip
import hashlib
import io
import json
import re
import os
import secrets
import time
import uuid
from datetime import datetime, timedelta

import psycopg2
import psycopg2.extras

import supplier

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Auth-Token',
    'Content-Type': 'application/json',
}

SESSION_DAYS = 7
XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

# Типы кузова автомобиля. Одна модель может выпускаться в нескольких —
# Rio бывает и седаном, и хэтчбеком, поэтому храним список.
BODY_TYPES = (
    'sedan', 'hatchback', 'liftback', 'wagon', 'crossover', 'suv', 'jeep',
    'pickup', 'minivan', 'van', 'coupe', 'cabrio',
)

# Как товар подбирается покупателю
FIT_MODES = ('vehicle', 'universal')

# Подбор проводки. Два РАЗНЫХ блока, и путать их нельзя:
#
#   WIRE_TECH  — с чем проводка умеет работать. Это фильтр: если у машины
#                штатный усилитель, а проводка на него не рассчитана, вариант
#                вообще не показываем.
#   WIRE_KEEPS — что останется работать у клиента после установки. Это не
#                фильтр, а объяснение цены: дешёвый переходник запустит
#                магнитолу, но климат с экрана пропадёт. Такой товар мы
#                показываем — но честно помечаем, что теряется.
WIRE_TECH = ('power', 'sound', 'wheel', 'amp', 'camera', 'can')
WIRE_KEEPS = ('climate', 'wheel', 'camera', 'amp', 'parktronic')

# yes — рассчитана на машины с этим, no — на машины без,
# any — параметр не влияет на совместимость (экономит вопрос покупателю)
WIRE_TECH_VALUES = ('yes', 'no', 'any')

# full — сохраняет всё нужное, basic — часть функций теряется,
# limited — существенные ограничения
WIRE_LEVELS = ('full', 'basic', 'limited')

# Сторона руля. Пусто — товар подходит и левому, и правому
WHEEL_SIDES = ('left', 'right')
WHEEL_RU = {'left': 'Левый', 'right': 'Правый'}
RU_WHEEL = {'левый': 'left', 'правый': 'right', 'левый руль': 'left', 'правый руль': 'right'}

# fixed — проводка для машины известна точно, вопросов не задаём
# select — вариантов несколько, уточняем по отмеченным параметрам
WIRE_MODES = ('fixed', 'select')


def clean_wire_tech(raw: dict) -> dict:
    """Оставляем только известные параметры с допустимыми значениями."""
    out = {}
    for key in WIRE_TECH:
        val = str((raw or {}).get(key) or '').strip().lower()
        if val in WIRE_TECH_VALUES:
            out[key] = val
    return out


def clean_wire_keeps(raw: dict) -> dict:
    """Сохраняемые функции — простые да/нет."""
    return {k: bool((raw or {}).get(k)) for k in WIRE_KEEPS if k in (raw or {})}


"""
Каталог перерос лимит ответа функции (4 МБ), поэтому большие ответы
отдаём сжатыми. Флаг ставит handler по заголовкам запроса.
"""
_GZIP = {'on': False}


def resp(status: int, payload: dict) -> dict:
    body = json.dumps(payload, ensure_ascii=False)

    if _GZIP['on'] and len(body) > 512_000:
        packed = gzip.compress(body.encode('utf-8'), 6)
        return {
            'statusCode': status,
            'headers': {**CORS, 'Content-Encoding': 'gzip'},
            'isBase64Encoded': True,
            'body': base64.b64encode(packed).decode('ascii'),
        }

    return {
        'statusCode': status,
        'headers': CORS,
        'isBase64Encoded': False,
        'body': body,
    }


def db():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def schema() -> str:
    return os.environ.get('MAIN_DB_SCHEMA', 'public')


def q(value) -> str:
    if value is None:
        return 'NULL'
    return "'" + str(value).replace("'", "''") + "'"


def qjson(value) -> str:
    return "'" + json.dumps(value, ensure_ascii=False).replace("'", "''") + "'::jsonb"


def qint(value, default=0) -> str:
    if value is None or value == '':
        return 'NULL'
    try:
        return str(int(value))
    except (TypeError, ValueError):
        return str(default)


def check_token(conn, token: str) -> bool:
    if not token:
        return False
    cur = conn.cursor()
    cur.execute(
        f"SELECT 1 FROM {schema()}.admin_sessions WHERE token = {q(token)} AND expires_at > NOW()"
    )
    ok = cur.fetchone() is not None
    cur.close()
    return ok


def _s3():
    # Загружаем тяжёлые библиотеки только когда они реально нужны:
    # иначе выгрузка каталога тратит секунды на импорт и упирается в таймаут
    import boto3

    return boto3.client(
        's3',
        endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    )


def upload_image(data_url: str) -> str:
    header, _, payload = data_url.partition(',')
    ext = 'jpg'
    if 'png' in header:
        ext = 'png'
    elif 'webp' in header:
        ext = 'webp'
    elif 'svg' in header:
        ext = 'svg'
    raw = base64.b64decode(payload)
    from image_optimizer import optimize

    body, ext, content_type = optimize(raw, ext)
    key = f"catalog/{uuid.uuid4().hex}.{ext}"
    s3 = _s3()
    s3.put_object(
        Bucket='files',
        Key=key,
        Body=body,
        ContentType=content_type,
        CacheControl='public, max-age=31536000, immutable',
    )
    return f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"


CDN_PREFIX = 'https://cdn.poehali.dev/projects/'

# Больше 15 МБ на одно фото — почти наверняка не товарный снимок
MAX_IMAGE_BYTES = 15 * 1024 * 1024

# Видео крупнее — это уже не ролик для карточки товара, а исходник
# со смартфона без сжатия. Оставляет запас на разумный обзор в 1-2 минуты.
MAX_VIDEO_BYTES = 180 * 1024 * 1024

VIDEO_CONTENT_TYPE = {
    'mp4': 'video/mp4',
    'webm': 'video/webm',
    'mov': 'video/quicktime',
    'mkv': 'video/x-matroska',
    'ogg': 'video/ogg',
    'avi': 'video/x-msvideo',
}


def _video_ext(name: str) -> str:
    tail = str(name or '').lower().rsplit('.', 1)
    ext = tail[1] if len(tail) == 2 else 'mp4'
    return ext if ext in VIDEO_CONTENT_TYPE else 'mp4'


# Сколько всего живёт функция. Платформа обрывает её на этом сроке,
# поэтому работу планируем так, чтобы успеть ответить самим.
FUNCTION_BUDGET = float(os.environ.get('FUNCTION_TIMEOUT', '2'))

# Запас на ответ и запись в базу — его не занимаем скачиванием
RESERVE = 0.6

# Сколько картинок качаем за один вызов. Качаем параллельно — сеть ждёт
# все сразу, — но каждую потом надо сжать и отправить в хранилище, а это
# уже процессор. Четыре укладываются в отведённые функции две секунды.
BATCH_IMAGES = 4


def _is_own_cdn(url: str) -> bool:
    key = os.environ.get('AWS_ACCESS_KEY_ID', '')
    return url.startswith(f'{CDN_PREFIX}{key}/bucket/')


def _cdn_key(url: str) -> str:
    key = os.environ.get('AWS_ACCESS_KEY_ID', '')
    return url.split(f'{CDN_PREFIX}{key}/bucket/', 1)[1].split('?')[0]


def reoptimize_url(url: str) -> str:
    """Скачивает картинку с CDN, пережимает в WebP и кладёт рядом новым файлом."""
    if not _is_own_cdn(url) or url.lower().split('?')[0].endswith(('.webp', '.svg')):
        return url

    src_key = _cdn_key(url)
    s3 = _s3()
    try:
        raw = s3.get_object(Bucket='files', Key=src_key)['Body'].read()
    except Exception:
        return url

    ext = src_key.rsplit('.', 1)[-1].lower()
    from image_optimizer import optimize

    data, new_ext, content_type = optimize(raw, ext)
    if new_ext != 'webp':
        return url

    new_key = f"catalog/{uuid.uuid4().hex}.webp"
    s3.put_object(
        Bucket='files',
        Key=new_key,
        Body=data,
        ContentType=content_type,
        CacheControl='public, max-age=31536000, immutable',
    )
    return f"{CDN_PREFIX}{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{new_key}"


def _is_external_image(url: str) -> bool:
    """Ссылка на картинку с чужого сайта, которую стоит перенести к нам."""
    u = str(url or '').strip()
    if not u.lower().startswith(('http://', 'https://')):
        return False
    return not u.startswith(CDN_PREFIX)


def download_external(url: str, timeout: float) -> tuple:
    """Только скачивает файл, ничего не обрабатывая.

    Вынесено отдельно, чтобы качать пачку картинок разом: сеть — самая
    долгая часть переноса, и пока один сайт думает, остальные грузятся.
    Возвращает (сырые_байты, расширение, причина_отказа).
    """
    import urllib.error
    import urllib.parse
    import urllib.request

    if timeout <= 0:
        return None, None, 'не хватило времени'

    # В адресах поставщика встречаются пробелы и кириллица («7 дюймов»).
    # Без кодирования библиотека такой адрес даже не принимает — и раньше
    # это выглядело как «сайт долго отвечает», из-за чего ссылка не попадала
    # в отказы и намертво запирала очередь переноса.
    try:
        parts = urllib.parse.urlsplit(url)
        safe = urllib.parse.urlunsplit((
            parts.scheme,
            parts.netloc,
            urllib.parse.quote(parts.path),
            urllib.parse.quote(parts.query, safe='=&'),
            '',
        ))
    except Exception:
        return None, None, 'неверный адрес картинки'

    req = urllib.request.Request(
        safe,
        headers={
            'User-Agent': 'Mozilla/5.0 (compatible; ShtatnoBot/1.0)',
            'Accept': 'image/*,*/*',
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            ctype = str(r.headers.get('Content-Type', '')).lower()
            raw = r.read(MAX_IMAGE_BYTES + 1)
    except urllib.error.HTTPError as e:
        return None, None, f'сайт ответил отказом ({e.code})'
    except (urllib.error.URLError, ValueError) as e:
        # Разделяем «сайт не успел» и «адрес битый»: первое стоит повторить,
        # второе — нет, иначе одна кривая ссылка блокирует всю очередь
        text = str(getattr(e, 'reason', e)).lower()
        if 'timed out' in text or 'timeout' in text:
            return None, None, 'сайт отвечает слишком долго'
        return None, None, 'не удалось открыть адрес'
    except Exception:
        return None, None, 'сайт отвечает слишком долго'

    if len(raw) > MAX_IMAGE_BYTES:
        return None, None, 'файл слишком большой'
    if not raw:
        return None, None, 'пустой файл'

    ext = 'jpg'
    for name in ('png', 'webp', 'gif', 'jpeg', 'jpg'):
        if name in ctype:
            ext = 'jpg' if name == 'jpeg' else name
            break
    else:
        tail = url.lower().split('?')[0].rsplit('.', 1)
        if len(tail) == 2 and tail[1] in ('png', 'webp', 'gif', 'jpg', 'jpeg'):
            ext = 'jpg' if tail[1] == 'jpeg' else tail[1]
        elif 'image' not in ctype:
            return None, None, 'по ссылке не картинка'

    return raw, ext, None


def store_image(raw: bytes, ext: str) -> tuple:
    """Сжимает скачанную картинку и кладёт в наше хранилище."""
    from image_optimizer import optimize

    try:
        data, new_ext, content_type = optimize(raw, ext)
    except Exception:
        return None, 'не удалось обработать картинку'

    key = f"catalog/{uuid.uuid4().hex}.{new_ext}"
    try:
        _s3().put_object(
            Bucket='files',
            Key=key,
            Body=data,
            ContentType=content_type,
            CacheControl='public, max-age=31536000, immutable',
        )
    except Exception:
        return None, 'не удалось сохранить у нас'

    return f"{CDN_PREFIX}{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}", None


def fetch_external_url(url: str, timeout: float = 0) -> tuple:
    """Скачивает чужую картинку, сжимает и кладёт к нам.

    Возвращает (наш_адрес, причина_отказа). Если не получилось — адрес None,
    а причину показываем в админке: пусть владелец каталога видит,
    какие именно ссылки не сработали и почему.

    timeout — сколько секунд готовы ждать чужой сайт. Ноль означает
    «сколько осталось от жизни функции».
    """
    import urllib.error
    import urllib.request

    req = urllib.request.Request(
        url,
        headers={
            # Часть сайтов не отдаёт файлы «неизвестным» клиентам
            'User-Agent': 'Mozilla/5.0 (compatible; ShtatnoBot/1.0)',
            'Accept': 'image/*,*/*',
        },
    )
    # Ждём чужой сайт ровно столько, сколько остаётся до лимита функции.
    # Ждать дольше бессмысленно: платформа оборвёт нас на середине
    # скачивания, картинка не сохранится, а попытка сгорит впустую.
    wait = timeout if timeout > 0 else FUNCTION_BUDGET - RESERVE
    if wait <= 0:
        return None, 'не хватило времени'

    try:
        with urllib.request.urlopen(req, timeout=wait) as r:
            ctype = str(r.headers.get('Content-Type', '')).lower()
            raw = r.read(MAX_IMAGE_BYTES + 1)
    except urllib.error.HTTPError as e:
        return None, f'сайт ответил отказом ({e.code})'
    except Exception:
        # Чаще всего сюда попадает медленный сайт, не уложившийся в срок.
        # Такую ссылку в чёрный список не заносим — попробуем ещё раз.
        return None, 'сайт отвечает слишком долго'

    if len(raw) > MAX_IMAGE_BYTES:
        return None, 'файл слишком большой'
    if not raw:
        return None, 'пустой файл'

    ext = 'jpg'
    for name in ('png', 'webp', 'gif', 'jpeg', 'jpg'):
        if name in ctype:
            ext = 'jpg' if name == 'jpeg' else name
            break
    else:
        tail = url.lower().split('?')[0].rsplit('.', 1)
        if len(tail) == 2 and tail[1] in ('png', 'webp', 'gif', 'jpg', 'jpeg'):
            ext = 'jpg' if tail[1] == 'jpeg' else tail[1]
        elif 'image' not in ctype:
            return None, 'по ссылке не картинка'

    from image_optimizer import optimize

    try:
        data, new_ext, content_type = optimize(raw, ext)
    except Exception:
        return None, 'не удалось обработать картинку'

    key = f"catalog/{uuid.uuid4().hex}.{new_ext}"
    try:
        _s3().put_object(
            Bucket='files',
            Key=key,
            Body=data,
            ContentType=content_type,
            CacheControl='public, max-age=31536000, immutable',
        )
    except Exception:
        return None, 'не удалось сохранить у нас'

    own = f"{CDN_PREFIX}{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"
    return own, None


DEFAULT_STOCK_NOTE = 'Под заказ - Отправка 1–3 дня'


def row_to_product(r: dict) -> dict:
    return {
        'id': r['id'],
        'slug': r['slug'],
        'sku': r.get('sku') or '',
        'name': r['name'],
        'category': r['category'],
        'price': r['price'],
        'oldPrice': r['old_price'],
        'proPrice': r.get('pro_price'),
        'ozonUrl': r.get('ozon_url') or '',
        'wbUrl': r.get('wb_url') or '',
        'install': r['install'],
        'warranty': r['warranty'],
        'yearFrom': r['year_from'],
        'yearTo': r['year_to'],
        'badge': r['badge'],
        'images': r['images'],
        'videoUrl': r.get('video_url') or '',
        'description': r['description'],
        'specs': r['specs'],
        'kit': r['kit'],
        'notes': r.get('notes') or [],
        'extra': r.get('extra') or [],
        'extraTitle': r.get('extra_title') or '',
        'fits': r['fits'],
        # Пусто — значит товар наследует умолчание своей категории
        'fitMode': r.get('fit_mode') or '',
        # Подбор проводки: с чем работает / что сохраняет / почему такая цена
        'wireTech': r.get('wire_tech') or {},
        'wireKeeps': r.get('wire_keeps') or {},
        'wireLevel': r.get('wire_level') or '',
        # Пусто — проводка встаёт на любой кузов
        'wireBodies': r.get('wire_bodies') or [],
        'wireWheel': r.get('wire_wheel') or '',
        'wireNote': r.get('wire_note') or '',
        # Проводки, подходящие к этой рамке. Пусто — ещё не размечено
        'frameWires': r.get('frame_wires') or [],
        # Проводка уже в коробке — отдельно предлагать не надо
        'wireIncluded': bool(r.get('wire_included')),
        'sortOrder': r['sort_order'],
        'popularity': r.get('popularity') or 0,
        'stock': r.get('stock_qty') or 0,
        'stockNote': r.get('stock_note') or DEFAULT_STOCK_NOTE,
        'isActive': r['is_active'],
    }


def row_to_order(r: dict) -> dict:
    return {
        'id': r['id'],
        'kind': r['kind'],
        'name': r['name'],
        'phone': r['phone'],
        'comment': r['comment'],
        'vehicle': r['vehicle'],
        'items': r['items'],
        'total': r['total'],
        'status': r['status'],
        'adminNote': r['admin_note'],
        'source': r['source'],
        'createdAt': r['created_at'].isoformat() if r['created_at'] else None,
    }


def row_to_guide(r: dict) -> dict:
    return {
        'id': r['id'],
        'slug': r['slug'],
        'title': r['title'],
        'excerpt': r['excerpt'],
        'cover': r['cover'],
        'duration': r['duration'],
        'difficulty': r['difficulty'],
        'tools': r['tools'],
        'blocks': r['blocks'],
        'sortOrder': r['sort_order'],
        'isActive': r['is_active'],
    }


SLUG_LIMIT = 60

TRANSLIT = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'h', 'ц': 'c', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch', 'ъ': '',
    'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
}

# Предлоги, союзы и «вода» — в адресе не нужны
SLUG_STOP = set("""
    dlya na s so i v vo k ko po ot do iz u o ob pri za pod nad a no ili zhe li by
    eto kak chto vse ves vsya vseh tip tipa vid vida goda godov god let
    shtuk sht komplekte takzhe ochen bolee samyy svoy nash vash lyuboy raznyh
    prochee drugoe novyy universalnyy
""".split())

# Перед числом предлог оставляем: «с 2016» и «до 2016» — разные товары
SLUG_KEEP_BEFORE_NUM = {'s', 'do', 'ot', 'po', 'pod', 'nad', 'iz'}


def slugify(text: str, limit: int = SLUG_LIMIT) -> str:
    """Название → короткий SEO-адрес на латинице, только значимые слова."""
    translit = ''.join(TRANSLIT.get(ch, ch) for ch in str(text or '').lower())
    words = [w for w in re.split(r'[^a-z0-9]+', translit) if w]

    kept = []
    seen = set()
    for i, w in enumerate(words):
        nxt = words[i + 1] if i + 1 < len(words) else ''
        keep_prefix = w in SLUG_KEEP_BEFORE_NUM and nxt.isdigit()
        if w in SLUG_STOP and not keep_prefix and kept:
            continue
        if w in seen and not w.isdigit():
            continue
        seen.add(w)
        kept.append(w)

    source = kept or words[:1]
    if not source:
        return 'tovar'

    slug = ''
    for w in source:
        candidate = w if not slug else slug + '-' + w
        if len(candidate) > limit:
            break
        slug = candidate
    return slug or source[0][:limit]


XLS_COLUMNS = [
    ('slug', 'Код (не менять)', 30),
    ('sku', 'Артикул', 16),
    ('name', 'Название', 46),
    ('category', 'Категория', 18),
    ('price', 'Цена', 12),
    ('oldPrice', 'Старая цена', 13),
    ('proPrice', 'Цена дилера', 13),
    ('warranty', 'Гарантия', 12),
    ('install', 'Установка', 18),
    ('ozonUrl', 'Ссылка Ozon', 34),
    ('wbUrl', 'Ссылка Wildberries', 34),
    ('yearFrom', 'Год с', 8),
    ('yearTo', 'Год по', 8),
    ('badge', 'Метка', 12),
    ('stock', 'Наличие (шт)', 13),
    ('stockNote', 'Если нет в наличии', 30),
    ('popularity', 'Популярность', 14),
    ('sortOrder', 'Порядок', 10),
    ('isActive', 'На сайте (да/нет)', 17),
    ('fits', 'Совместимость', 60),
    ('images', 'Фото (ссылки через ; — можно с других сайтов)', 46),
    ('description', 'Описание (абзацы через |)', 50),
    ('specs', 'Характеристики (Имя=Значение;)', 50),
    ('kit', 'Комплектация (через ;)', 40),
]


def fits_to_text(fits: dict) -> str:
    return ' | '.join(
        f"{brand}: {', '.join(models)}" for brand, models in (fits or {}).items()
    )


def text_to_fits(text: str) -> dict:
    out: dict = {}
    for chunk in str(text or '').split('|'):
        if ':' not in chunk:
            continue
        brand, _, models = chunk.partition(':')
        brand = brand.strip()
        items = [m.strip() for m in models.split(',') if m.strip()]
        if brand and items:
            out[brand] = items
    return out


def specs_to_text(specs: list) -> str:
    parts = []
    for row in specs or []:
        if isinstance(row, (list, tuple)) and len(row) >= 2:
            parts.append(f"{row[0]}={row[1]}")
    return '; '.join(parts)


def text_to_specs(text: str) -> list:
    out = []
    for chunk in str(text or '').split(';'):
        if '=' not in chunk:
            continue
        k, _, v = chunk.partition('=')
        if k.strip():
            out.append([k.strip(), v.strip()])
    return out


def list_to_text(items: list, sep: str = '; ') -> str:
    return sep.join(str(i) for i in (items or []))


def text_to_list(text: str, sep: str) -> list:
    return [p.strip() for p in str(text or '').split(sep) if p.strip()]


def build_xlsx(products: list, brands: list, categories: list = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color='FFFFFF', size=10)
    head_fill = PatternFill('solid', fgColor='1B1B1B')
    lock_fill = PatternFill('solid', fgColor='E01B0C')

    ws = wb.active
    ws.title = 'Товары'
    for i, (key, title, width) in enumerate(XLS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.font = head_font
        cell.fill = lock_fill if key == 'slug' else head_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'C2'

    for r, p in enumerate(products, start=2):
        values = {
            'slug': p.get('slug', ''),
            'sku': p.get('sku', ''),
            'name': p.get('name', ''),
            'category': p.get('category', ''),
            'price': p.get('price', 0),
            'oldPrice': p.get('oldPrice'),
            'proPrice': p.get('proPrice'),
            'warranty': p.get('warranty', ''),
            'install': p.get('install', ''),
            'ozonUrl': p.get('ozonUrl', ''),
            'wbUrl': p.get('wbUrl', ''),
            'yearFrom': p.get('yearFrom', ''),
            'yearTo': p.get('yearTo', ''),
            'badge': p.get('badge') or '',
            'stock': p.get('stock', 0),
            'stockNote': p.get('stockNote') or DEFAULT_STOCK_NOTE,
            'popularity': p.get('popularity', 0),
            'sortOrder': p.get('sortOrder', 100),
            'isActive': 'да' if p.get('isActive', True) else 'нет',
            'fits': fits_to_text(p.get('fits') or {}),
            'images': list_to_text(p.get('images') or []),
            'description': ' | '.join(p.get('description') or []),
            'specs': specs_to_text(p.get('specs') or []),
            'kit': list_to_text(p.get('kit') or []),
        }
        for i, (key, _t, _w) in enumerate(XLS_COLUMNS, start=1):
            ws.cell(row=r, column=i, value=values.get(key))

    wsb = wb.create_sheet('Марки и модели')
    for i, title in enumerate(['Марка', 'Модели (через запятую)'], start=1):
        cell = wsb.cell(row=1, column=i, value=title)
        cell.font = head_font
        cell.fill = head_fill
    wsb.column_dimensions['A'].width = 22
    wsb.column_dimensions['B'].width = 80
    wsb.freeze_panes = 'A2'
    for r, b in enumerate(brands, start=2):
        wsb.cell(row=r, column=1, value=b.get('name', ''))
        wsb.cell(row=r, column=2, value=', '.join(b.get('models') or []))

    wsc = wb.create_sheet('Категории')
    for i, title in enumerate(
        ['Категория', 'Порядок', 'Характеристики категории (через запятую)'], start=1
    ):
        cell = wsc.cell(row=1, column=i, value=title)
        cell.font = head_font
        cell.fill = head_fill
    wsc.column_dimensions['A'].width = 30
    wsc.column_dimensions['B'].width = 10
    wsc.column_dimensions['C'].width = 60
    wsc.freeze_panes = 'A2'
    for r, c in enumerate(categories or [], start=2):
        wsc.cell(row=r, column=1, value=c.get('name', ''))
        wsc.cell(row=r, column=2, value=c.get('sortOrder', (r - 1) * 10))
        wsc.cell(row=r, column=3, value=', '.join(c.get('specFields') or []))

    wsh = wb.create_sheet('Как заполнять')
    tips = [
        ['Подсказка', 'Что важно знать'],
        ['Столбец «Код»', 'Не меняйте и не удаляйте — по нему товар находится при загрузке. Для нового товара оставьте пустым.'],
        ['Новый товар', 'Добавьте строку внизу, заполните название, категорию и цену. Код создастся сам.'],
        ['Цены', 'Числа. Пробелы, запятую и знак рубля уберём сами: «15 900,00 ₽» прочитается как 15900. Пустая «Старая цена» — скидки нет.'],
        ['Установка', 'Куда ставится: «в штатное место», «в ручку багажника». Показывается в карточке.'],
        ['Ozon и Wildberries', 'Ссылки на товар на маркетплейсах. Пусто — кнопка не показывается.'],
        ['На сайте', '«да» — товар виден покупателям, «нет» — скрыт.'],
        ['Совместимость', 'Формат: Lada: Vesta SW Cross, Granta | Kia: Rio, Seltos'],
        ['Фото', 'Ссылки на картинки через точку с запятой. Можно вставлять '
                 'адреса с других сайтов — после загрузки файла нажмите '
                 '«Перенести фото к нам» в разделе настроек.'],
        ['Описание', 'Абзацы разделяйте вертикальной чертой |. Внутри самого текста черту не используйте — она разорвёт абзац.'],
        ['Характеристики', 'Формат: Гарантия=5 лет; Материал=сталь 2 мм'],
        ['Комплектация', 'Пункты через точку с запятой.'],
        ['Марки и модели', 'Отдельный лист. Модели одной марки — через запятую.'],
        ['Категории', 'Отдельный лист. Порядок задаёт вид фильтра на сайте. Характеристики категории показываются в каталоге под товаром.'],
        ['Загрузка', 'Ничего не удаляется: совпавшие по коду товары обновятся, новые добавятся.'],
        ['Лишние столбцы', 'Можно удалить ненужные столбцы — эти поля у товара останутся прежними. Столбцы «Код», «Название», «Категория» нужны всегда.'],
    ]
    for r, row in enumerate(tips, start=1):
        for c, val in enumerate(row, start=1):
            cell = wsh.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = head_font
                cell.fill = head_fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    wsh.column_dimensions['A'].width = 26
    wsh.column_dimensions['B'].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_xlsx(data: bytes) -> tuple:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    products = []
    brands = []
    categories = []

    titles = {t: k for k, t, _w in XLS_COLUMNS}

    if 'Товары' in wb.sheetnames:
        ws = wb['Товары']
    else:
        ws = wb.worksheets[0]

    header = [str(c.value or '').strip() for c in ws[1]]
    index = {}
    for i, title in enumerate(header):
        key = titles.get(title)
        if key:
            index[key] = i

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, '') for v in row):
            continue

        def get(key, default=''):
            i = index.get(key)
            if i is None or i >= len(row):
                return default
            v = row[i]
            return default if v is None else v

        name = str(get('name')).strip()
        if not name:
            continue

        def as_int(key, default=None):
            # Из 1С и Excel цены приходят как «15 900,00» или «15 900.00 ₽»
            raw = str(get(key, '')).strip()
            raw = raw.replace(' ', '').replace('\u00a0', '').replace(',', '.')
            raw = re.sub(r'[^0-9.\-]', '', raw)
            if raw in ('', '-', '.'):
                return default
            try:
                return int(round(float(raw)))
            except ValueError:
                return default

        active_raw = str(get('isActive', 'да')).strip().lower()
        products.append({
            'slug': str(get('slug')).strip(),
            'sku': str(get('sku')).strip(),
            'name': name,
            'category': str(get('category')).strip(),
            'price': as_int('price', 0) or 0,
            'oldPrice': as_int('oldPrice'),
            'proPrice': as_int('proPrice'),
            'warranty': str(get('warranty')).strip(),
            'install': str(get('install')).strip(),
            'ozonUrl': str(get('ozonUrl')).strip(),
            'wbUrl': str(get('wbUrl')).strip(),
            'yearFrom': as_int('yearFrom', 2010),
            'yearTo': as_int('yearTo', 2026),
            'badge': str(get('badge')).strip() or None,
            'stock': as_int('stock', 0) or 0,
            'stockNote': str(get('stockNote')).strip() or DEFAULT_STOCK_NOTE,
            'popularity': as_int('popularity', 0) or 0,
            'sortOrder': as_int('sortOrder', 100) or 100,
            'isActive': active_raw not in ('нет', 'no', 'false', '0', 'скрыт'),
            'fits': text_to_fits(get('fits')),
            'images': text_to_list(get('images'), ';'),
            'description': text_to_list(get('description'), '|'),
            'specs': text_to_specs(get('specs')),
            'kit': text_to_list(get('kit'), ';'),
            # Какие столбцы реально были в файле — остальные поля не трогаем
            '_columns': sorted(index.keys()),
        })

    if 'Марки и модели' in wb.sheetnames:
        wsb = wb['Марки и модели']
        for row in wsb.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            models = str(row[1] or '')
            brands.append({
                'name': str(row[0]).strip(),
                'models': [m.strip() for m in models.split(',') if m.strip()],
            })

    if 'Категории' in wb.sheetnames:
        wsc = wb['Категории']
        for row in wsc.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            fields = str(row[2] or '') if len(row) > 2 else ''
            try:
                order = int(row[1]) if len(row) > 1 and row[1] is not None else 0
            except (TypeError, ValueError):
                order = 0
            categories.append({
                'name': str(row[0]).strip(),
                'sortOrder': order,
                'specFields': [f.strip() for f in fields.split(',') if f.strip()],
            })

    return products, brands, categories


def build_brands_xlsx(brands: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color='FFFFFF', size=10)
    head_fill = PatternFill('solid', fgColor='1B1B1B')

    ws = wb.active
    ws.title = 'Марки и модели'
    for i, (title, width) in enumerate(
        [('Марка', 26), ('Модель', 40), ('Порядок марки', 15)], start=1
    ):
        cell = ws.cell(row=1, column=i, value=title)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'

    row = 2
    for i, b in enumerate(brands or []):
        bname = str(b.get('name', '')).strip()
        if not bname:
            continue
        models = [str(m).strip() for m in (b.get('models') or []) if str(m).strip()]
        if not models:
            ws.cell(row=row, column=1, value=bname)
            ws.cell(row=row, column=3, value=(i + 1) * 10)
            row += 1
            continue
        for m in models:
            ws.cell(row=row, column=1, value=bname)
            ws.cell(row=row, column=2, value=m)
            ws.cell(row=row, column=3, value=(i + 1) * 10)
            row += 1

    wsh = wb.create_sheet('Как заполнять')
    tips = [
        ['Подсказка', 'Что важно знать'],
        ['Одна строка', 'Одна строка — одна пара «марка + модель». Марка повторяется столько раз, сколько у неё моделей.'],
        ['Новая марка', 'Просто добавьте строки внизу: впишите название марки и модель. Марка создастся сама.'],
        ['Новая модель', 'Добавьте строку с уже существующей маркой и новым названием модели.'],
        ['Переименование', 'Меняйте название прямо в ячейке. Учтите: у товаров совместимость привязана к старому названию.'],
        ['Порядок марки', 'Число, по которому марки сортируются в подборе. Меньше — выше. Можно оставить пустым.'],
        ['Марка без моделей', 'Оставьте столбец «Модель» пустым — марка появится в списке без моделей.'],
        ['Загрузка', 'Марки из файла заменяют список целиком: чего нет в файле — того не будет на сайте.'],
        ['Дубли', 'Одинаковые пары «марка + модель» схлопываются автоматически, можно не следить.'],
    ]
    for r, line in enumerate(tips, start=1):
        for c, val in enumerate(line, start=1):
            cell = wsh.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = head_font
                cell.fill = head_fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    wsh.column_dimensions['A'].width = 26
    wsh.column_dimensions['B'].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


WIRES_CATEGORY = 'Переходники для подключения магнитол'
FRAMES_CATEGORY = 'Переходные рамки для магнитол'

# Подписи колонок файла подбора. Меняешь текст — поправь и разбор при импорте
WIRE_TECH_TITLES = [
    ('power', 'Питание'),
    ('sound', 'Акустика'),
    ('wheel', 'Кнопки руля'),
    ('amp', 'Усилитель'),
    ('camera', 'Камера'),
    ('can', 'CAN'),
]
WIRE_KEEPS_TITLES = [
    ('climate', 'Сохраняет климат'),
    ('wheel', 'Сохраняет кнопки'),
    ('camera', 'Сохраняет камеру'),
    ('amp', 'Сохраняет усилитель'),
    ('parktronic', 'Сохраняет парктроники'),
]
BODY_RU = {
    'sedan': 'Седан',
    'hatchback': 'Хэтчбек',
    'liftback': 'Лифтбек',
    'wagon': 'Универсал',
    'crossover': 'Кроссовер',
    'suv': 'Внедорожник',
    'jeep': 'Джип',
    'pickup': 'Пикап',
    'minivan': 'Минивэн',
    'van': 'Фургон',
    'coupe': 'Купе',
    'cabrio': 'Кабриолет',
}
RU_BODY = {v.lower(): k for k, v in BODY_RU.items()}
# Живые написания: «хэтчбэк», «хетчбек», «джип» — человек пишет как привык,
# а файл из-за одной буквы ругаться не должен
RU_BODY.update(
    {
        'хэтчбэк': 'hatchback',
        'хетчбек': 'hatchback',
        'хетчбэк': 'hatchback',
        'хэтчбек 5 дв': 'hatchback',
        'хэтчбэк 5 дв': 'hatchback',
        'лифтбэк': 'liftback',
        'вагон': 'wagon',
        'паркетник': 'crossover',
        'внедорожник/кроссовер': 'crossover',
        'рамный внедорожник': 'jeep',
        'микроавтобус': 'minivan',
        'минивен': 'minivan',
        'вэн': 'van',
        'кабрио': 'cabrio',
    }
)

LEVEL_RU = {'full': 'Полная', 'basic': 'Базовая', 'limited': 'Ограниченная'}
RU_LEVEL = {v.lower(): k for k, v in LEVEL_RU.items()}
TECH_RU = {'yes': 'Да', 'no': 'Нет', 'any': 'Неважно'}
RU_TECH = {'да': 'yes', 'нет': 'no', 'неважно': 'any'}
MODE_RU = {'fixed': 'Фиксированный', 'select': 'Подбор'}
RU_MODE = {v.lower(): k for k, v in MODE_RU.items()}


def vehicle_rows(products: list, brands: list, scope: str) -> list:
    """
    Строки листа «Подбор проводки»: одна машина = одна строка.

    Размечать все полторы тысячи моделей смысла нет. Комплект собирается
    только там, где под машину есть и рамка, и проводка — без рамки
    магнитоле некуда встать. А деньги теряются там, где к тому же
    несколько проводок с большим разбросом цен: человек берёт дешёвую,
    теряет штатную функцию и уходит. Такие машины поднимаем наверх.
    """
    wires, frames = {}, set()
    for p in products:
        cat = p.get('category') or ''
        for brand, models in (p.get('fits') or {}).items():
            for model in models or []:
                key = (brand, model)
                if cat == WIRES_CATEGORY:
                    wires.setdefault(key, []).append(p)
                elif cat == FRAMES_CATEGORY:
                    frames.add(key)

    known = set()
    for b in brands:
        for m in b.get('models') or []:
            known.add((b['name'], m))

    rows = []
    for key in sorted(known):
        ws = wires.get(key) or []
        has_frame = key in frames
        prices = [p.get('price') or 0 for p in ws if p.get('price')]
        spread = (max(prices) - min(prices)) if len(prices) > 1 else 0
        hot = (
            has_frame
            and len(ws) > 1
            and spread > 1000
            and max(prices) >= 3 * max(min(prices), 1)
        )
        if scope == 'hot' and not hot:
            continue
        if scope == 'kit' and not (has_frame and ws):
            continue
        if scope != 'all' and not ws:
            continue
        rows.append(
            {
                'brand': key[0],
                'model': key[1],
                'wires': len(ws),
                'spread': spread,
                'hot': hot,
            }
        )
    # Сверху — где разброс цен больше: там сделки и срываются
    rows.sort(key=lambda r: (-r['spread'], r['brand'], r['model']))
    return rows


def build_wiring_xlsx(products: list, brands: list, saved: list, scope: str) -> bytes:
    """Файл для массовой разметки: проводки и настройки машин."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    white = Font(bold=True, color='FFFFFF', size=10)
    dark = PatternFill('solid', fgColor='1B1B1B')
    red = PatternFill('solid', fgColor='C8102E')
    grey = PatternFill('solid', fgColor='6B7280')
    mid = Alignment(vertical='center', wrap_text=True, horizontal='center')

    # ---------- Лист 1: проводки ----------
    ws = wb.active
    ws.title = 'Проводки'
    base = [
        ('Код (не менять)', 34),
        ('Название', 52),
        ('Цена', 10),
        ('Год с', 8),
        ('Год по', 8),
    ]
    tech = [(t, 12) for _k, t in WIRE_TECH_TITLES]
    keeps = [
        ('Кузов (пусто — любой)', 22),
        ('Руль (пусто — любой)', 20),
        ('Уровень', 14),
    ] + [
        (t, 17) for _k, t in WIRE_KEEPS_TITLES
    ]
    tail = [('Описание совместимости (видит покупатель)', 60)]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(base))
    c = ws.cell(1, 1, 'Уже есть — не трогаете')
    c.fill, c.font, c.alignment = grey, white, mid
    p1 = len(base) + 1
    ws.merge_cells(start_row=1, start_column=p1, end_row=1, end_column=p1 + len(tech) - 1)
    c = ws.cell(1, p1, 'С чем работает — для каких машин проводка')
    c.fill, c.font, c.alignment = red, white, mid
    p2 = p1 + len(tech)
    ws.merge_cells(
        start_row=1, start_column=p2, end_row=1, end_column=p2 + len(keeps) + len(tail) - 1
    )
    c = ws.cell(1, p2, 'Что останется работать у клиента')
    c.fill, c.font, c.alignment = red, white, mid

    cols = base + tech + keeps + tail
    for i, (title, width) in enumerate(cols, start=1):
        cell = ws.cell(2, i, title)
        cell.font, cell.fill, cell.alignment = white, dark, mid
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 34
    ws.freeze_panes = 'C3'

    wires = [p for p in products if (p.get('category') or '') == WIRES_CATEGORY]
    for r, p in enumerate(wires, start=3):
        t = p.get('wireTech') or {}
        k = p.get('wireKeeps') or {}
        vals = [
            p.get('slug', ''),
            p.get('name', ''),
            p.get('price', 0),
            p.get('yearFrom', ''),
            p.get('yearTo', ''),
        ]
        vals += [TECH_RU.get(t.get(key, ''), '') for key, _t in WIRE_TECH_TITLES]
        vals += [', '.join(BODY_RU.get(b, b) for b in (p.get('wireBodies') or []))]
        vals += [WHEEL_RU.get(p.get('wireWheel') or '', '')]
        vals += [LEVEL_RU.get(p.get('wireLevel') or '', '')]
        vals += [
            ('Да' if k[key] else 'Нет') if key in k else ''
            for key, _t in WIRE_KEEPS_TITLES
        ]
        vals += [p.get('wireNote') or '']
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(r, i, v)
            cell.alignment = Alignment(vertical='top', wrap_text=i > len(base))

    last = max(len(wires) + 3, 400)
    dv = DataValidation(type='list', formula1='"Да,Нет,Неважно"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'F3:{get_column_letter(p1 + len(tech) - 1)}{last}')
    # Один кузов выбирается из списка. Несколько (редкий случай) — руками
    # через запятую, поэтому запрет на ввод своего значения не ставим
    dv_body = DataValidation(
        type='list',
        formula1='"' + ','.join(BODY_RU.values()) + '"',
        allow_blank=True,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv_body)
    dv_body.add(f'{get_column_letter(p2)}3:{get_column_letter(p2)}{last}')

    dv2 = DataValidation(
        type='list', formula1='"Полная,Базовая,Ограниченная"', allow_blank=True
    )
    ws.add_data_validation(dv2)
    dv_wheel = DataValidation(
        type='list', formula1='"Левый,Правый"', allow_blank=True
    )
    ws.add_data_validation(dv_wheel)
    dv_wheel.add(f'{get_column_letter(p2 + 1)}3:{get_column_letter(p2 + 1)}{last}')

    dv2.add(f'{get_column_letter(p2 + 2)}3:{get_column_letter(p2 + 2)}{last}')
    dv3 = DataValidation(type='list', formula1='"Да,Нет"', allow_blank=True)
    ws.add_data_validation(dv3)
    dv3.add(
        f'{get_column_letter(p2 + 3)}3:'
        f'{get_column_letter(p2 + 2 + len(WIRE_KEEPS_TITLES))}{last}'
    )

    # ---------- Лист 2: машины ----------
    w2 = wb.create_sheet('Подбор проводки')
    vcols = [
        ('Проводок', 10),
        ('Разброс цен', 13),
        ('Марка', 16),
        ('Модель', 20),
        ('Год от', 9),
        ('Год по', 9),
        ('Кузов (пусто — любой)', 22),
        ('Руль (пусто — любой)', 18),
        ('Тип подбора', 16),
        ('Код проводки (для «Фиксированный»)', 34),
        ('Обоснование — увидит покупатель', 46),
        ('Спросить про усилитель', 13),
        ('Спросить про камеру', 13),
        ('Спросить про CAN', 13),
    ]
    w2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(vcols))
    c = w2.cell(1, 1, 'Сверху машины с самым большим разбросом цен — начинайте с них')
    c.fill, c.font, c.alignment = red, white, mid
    for i, (title, width) in enumerate(vcols, start=1):
        cell = w2.cell(2, i, title)
        cell.font, cell.fill, cell.alignment = white, dark, mid
        w2.column_dimensions[get_column_letter(i)].width = width
    w2.row_dimensions[2].height = 34
    w2.freeze_panes = 'E3'

    # У модели бывает несколько поколений — выгружаем каждое своей строкой
    have = {}
    for v in saved:
        have.setdefault((v['brand'], v['model']), []).append(v)

    r = 3
    for row in vehicle_rows(products, brands, scope):
        gens = have.get((row['brand'], row['model'])) or [{}]
        for s_row in gens:
            ask = s_row.get('ask') or {}
            vals = [
                row['wires'],
                row['spread'] or '',
                row['brand'],
                row['model'],
                s_row.get('yearFrom') or '',
                s_row.get('yearTo') if (s_row.get('yearTo') or 2100) < 2100 else '',
                ', '.join(
                    BODY_RU.get(b, b) for b in (s_row.get('bodies') or [])
                ),
                WHEEL_RU.get(s_row.get('wheel') or '', ''),
                MODE_RU.get(s_row.get('mode') or '', ''),
                s_row.get('wireSlug') or '',
                s_row.get('reason') or '',
                'Да' if ask.get('amp') else '',
                'Да' if ask.get('camera') else '',
                'Да' if ask.get('can') else '',
            ]
            for i, v in enumerate(vals, start=1):
                cell = w2.cell(r, i, v)
                cell.alignment = Alignment(vertical='top', wrap_text=i in (4, 11))
                if i <= 2:
                    cell.font = Font(color='888888', size=9)
            r += 1

    dv4 = DataValidation(
        type='list', formula1='"Фиксированный,Подбор"', allow_blank=True
    )
    w2.add_data_validation(dv4)
    dv4.add('I3:I3000')
    dv5 = DataValidation(type='list', formula1='"Да,Нет"', allow_blank=True)
    w2.add_data_validation(dv5)
    dv5.add('L3:N3000')
    dv_b = DataValidation(
        type='list',
        formula1='"' + ','.join(BODY_RU.values()) + '"',
        allow_blank=True,
        showErrorMessage=False,
    )
    w2.add_data_validation(dv_b)
    dv_b.add('G3:G3000')
    dv_w = DataValidation(
        type='list', formula1='"Левый,Правый"', allow_blank=True
    )
    w2.add_data_validation(dv_w)
    dv_w.add('H3:H3000')

    # ---------- Лист 3: подсказки ----------
    w3 = wb.create_sheet('Как заполнять')
    tips = [
        ('ГЛАВНОЕ', ''),
        (
            'Два разных блока',
            'На листе «Проводки» слева — «С чем работает»: для каких машин проводка '
            'предназначена. Если не подходит — покупатель её вообще не увидит. '
            'Справа — «Что останется работать»: товар показываем, но честно пишем, '
            'что теряется. Не путайте эти блоки.',
        ),
        (
            'Неважно',
            'Ставьте, когда параметр не влияет на совместимость. Каждое «Неважно» — '
            'это вопрос, который не придётся задавать покупателю.',
        ),
        (
            'Кузов',
            'Заполняйте, только если проводка встаёт не на все кузова. Пусто — '
            'подходит любому. Несколько кузовов пишите через запятую: '
            '«Седан, Универсал». У машин кузова уже размечены, так что система '
            'сама поймёт, когда спросить покупателя.',
        ),
        (
            'Уровень',
            'Полная — сохраняет всё нужное. Базовая — магнитола работает, часть '
            'функций теряется. Ограниченная — существенные ограничения.',
        ),
        (
            'Описание совместимости',
            'Этот текст УВИДИТ ПОКУПАТЕЛЬ. Пишите по-человечески: «магнитола '
            'заработает, но климат на экране пропадёт».',
        ),
        ('', ''),
        ('ЛИСТ «ПОДБОР ПРОВОДКИ»', ''),
        (
            'Сортировка',
            'Сверху машины с самым большим разбросом цен между дешёвой и дорогой '
            'проводкой — там клиенты и уходят. Идите сверху вниз, можно остановиться '
            'в любой момент.',
        ),
        (
            'Фиксированный',
            'Знаете точную проводку — впишите её код с листа «Проводки». Вопросов '
            'покупателю не будет.',
        ),
        (
            'Подбор',
            'Вариантов несколько. Отметьте «Да» только у тех параметров, которые '
            'ДЕЙСТВИТЕЛЬНО решают, какую проводку брать.',
        ),
        (
            'Важно про галки',
            'Галка не значит «в машине есть камера». Галка значит «от камеры '
            'зависит, какая проводка нужна».',
        ),
        (
            'Несколько поколений',
            'У модели может быть несколько строк: Civic 2006-2011 хэтчбек и '
            'Civic 2012-2020 седан. Просто добавьте вторую строку с той же '
            'маркой и моделью — она не затрёт первую.',
        ),
        (
            'Кузов и руль',
            'Заполняйте, только если для этого поколения проводка своя. Пусто '
            'значит любой. По кузову система различает поколения: владельцу '
            'седана не покажет проводку для хэтчбека.',
        ),
        (
            'Пустые строки',
            'Незаполненные строки просто пропускаются. Заполняйте столько, '
            'сколько успеете.',
        ),
    ]
    for r, (a, b) in enumerate(tips, start=1):
        ca = w3.cell(r, 1, a)
        cb = w3.cell(r, 2, b)
        cb.alignment = Alignment(vertical='top', wrap_text=True)
        if a and not b:
            ca.fill, ca.font = dark, white
        else:
            ca.font = Font(bold=True, size=10)
        ca.alignment = Alignment(vertical='top')
        w3.row_dimensions[r].height = 42 if b else 22
    w3.column_dimensions['A'].width = 28
    w3.column_dimensions['B'].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_wiring_xlsx(data: bytes) -> dict:
    """
    Разбираем заполненный файл подбора.

    Строки с ошибками не берём молча — возвращаем список проблем с номером
    строки, чтобы человек понял, что именно поправить.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    wires, cars, problems = [], [], []

    def cell(row, idx):
        return str(row[idx].value or '').strip() if idx < len(row) else ''

    if 'Проводки' in wb.sheetnames:
        ws = wb['Проводки']
        n_base = 5
        for r, row in enumerate(ws.iter_rows(min_row=3), start=3):
            slug = cell(row, 0)
            if not slug:
                continue
            tech = {}
            for i, (key, _t) in enumerate(WIRE_TECH_TITLES):
                v = cell(row, n_base + i).lower()
                if v and v in RU_TECH:
                    tech[key] = RU_TECH[v]
                elif v:
                    problems.append(f'Проводки, строка {r}: «{v}» — пишите Да/Нет/Неважно')
            p2 = n_base + len(WIRE_TECH_TITLES)
            # Кузова через запятую. Пусто — проводка встаёт на любой
            bodies = []
            for part in cell(row, p2).split(','):
                name = part.strip().lower()
                if not name:
                    continue
                if name in RU_BODY:
                    bodies.append(RU_BODY[name])
                else:
                    problems.append(
                        f'Проводки, строка {r}: кузов «{part.strip()}» неизвестен'
                    )
            wheel_raw = cell(row, p2 + 1).lower()
            wheel = RU_WHEEL.get(wheel_raw, '')
            if wheel_raw and not wheel:
                problems.append(
                    f'Проводки, строка {r}: руль «{wheel_raw}» — пишите Левый или Правый'
                )
            lvl_raw = cell(row, p2 + 2).lower()
            level = RU_LEVEL.get(lvl_raw, '')
            if lvl_raw and not level:
                problems.append(
                    f'Проводки, строка {r}: уровень «{lvl_raw}» — '
                    f'пишите Полная/Базовая/Ограниченная'
                )
            keeps = {}
            for i, (key, _t) in enumerate(WIRE_KEEPS_TITLES):
                v = cell(row, p2 + 3 + i).lower()
                if v in ('да', 'нет'):
                    keeps[key] = v == 'да'
                elif v:
                    problems.append(f'Проводки, строка {r}: «{v}» — пишите Да или Нет')
            wires.append(
                {
                    'slug': slug,
                    'tech': tech,
                    'keeps': keeps,
                    'level': level,
                    'bodies': bodies,
                    'wheel': wheel,
                    'note': cell(row, p2 + 3 + len(WIRE_KEEPS_TITLES))[:600],
                }
            )

    if 'Подбор проводки' in wb.sheetnames:
        ws = wb['Подбор проводки']
        for r, row in enumerate(ws.iter_rows(min_row=3), start=3):
            brand, model = cell(row, 2), cell(row, 3)
            mode_raw = cell(row, 8).lower()
            if not brand or not model or not mode_raw:
                continue
            mode = RU_MODE.get(mode_raw, '')
            if not mode:
                problems.append(
                    f'Подбор, строка {r}: тип «{mode_raw}» — '
                    f'пишите Фиксированный или Подбор'
                )
                continue

            def year(idx, default):
                raw = cell(row, idx)
                if not raw:
                    return default
                try:
                    return int(float(raw))
                except ValueError:
                    problems.append(f'Подбор, строка {r}: год «{raw}» — нужно число')
                    return default

            # Кузова поколения: «Хэтчбек» или «Седан, Универсал»
            bodies = []
            for part in cell(row, 6).split(','):
                name = part.strip().lower()
                if not name:
                    continue
                if name in RU_BODY:
                    bodies.append(RU_BODY[name])
                else:
                    problems.append(
                        f'Подбор, строка {r}: кузов «{part.strip()}» неизвестен'
                    )

            wheel_raw = cell(row, 7).lower()
            wheel = RU_WHEEL.get(wheel_raw, '')
            if wheel_raw and not wheel:
                problems.append(
                    f'Подбор, строка {r}: руль «{wheel_raw}» — '
                    f'пишите Левый или Правый'
                )

            wire_slug = cell(row, 9)
            if mode == 'fixed' and not wire_slug:
                problems.append(
                    f'Подбор, строка {r}: для «Фиксированный» нужен код проводки'
                )
                continue
            cars.append(
                {
                    'brand': brand,
                    'model': model,
                    'yearFrom': year(4, 1990),
                    'yearTo': year(5, 2100),
                    'bodies': bodies,
                    'wheel': wheel,
                    'mode': mode,
                    'wireSlug': wire_slug,
                    'reason': cell(row, 10)[:600],
                    'ask': {
                        'amp': cell(row, 11).lower() == 'да',
                        'camera': cell(row, 12).lower() == 'да',
                        'can': cell(row, 13).lower() == 'да',
                    },
                }
            )

    return {'wires': wires, 'cars': cars, 'problems': problems}


def parse_brands_xlsx(data: bytes) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb['Марки и модели'] if 'Марки и модели' in wb.sheetnames else wb.worksheets[0]

    header = [str(c.value or '').strip().lower() for c in ws[1]]
    wide = 'модели (через запятую)' in header

    order: dict = {}
    models_by_brand: dict = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        bname = str(row[0]).strip()
        if not bname:
            continue
        raw = str(row[1] or '') if len(row) > 1 else ''
        items = (
            [m.strip() for m in raw.split(',') if m.strip()]
            if wide
            else ([raw.strip()] if raw.strip() else [])
        )
        bucket = models_by_brand.setdefault(bname, [])
        for m in items:
            if m not in bucket:
                bucket.append(m)
        if bname not in order:
            sort_raw = str(row[2] or '').strip() if len(row) > 2 else ''
            try:
                order[bname] = int(float(sort_raw)) if sort_raw else len(order) * 10 + 10
            except ValueError:
                order[bname] = len(order) * 10 + 10

    return [
        {'name': name, 'models': models_by_brand[name]}
        for name in sorted(models_by_brand, key=lambda n: (order.get(n, 999), n))
    ]


def replace_brands(conn, brands: list) -> int:
    """
    Заменяет справочник марок целиком (загрузка из Excel).

    Типы кузова в таблицу не попадают, поэтому запоминаем их до удаления
    и возвращаем тем моделям, что остались. Иначе загрузка обновлённого
    списка марок молча стирала бы всю разметку кузовов.
    """
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT name, model_bodies FROM {schema()}.brands")
    kept = {r['name']: (r['model_bodies'] or {}) for r in cur.fetchall()}
    cur.close()

    cur = conn.cursor()
    cur.execute(f"DELETE FROM {schema()}.brands")

    # Все марки одной вставкой: раньше на каждую шёл отдельный запрос,
    # и на полном справочнике (60+ марок) сохранение не укладывалось
    # в отведённое время — админка отвечала ошибкой
    values = []
    for i, b in enumerate(brands):
        name = str(b.get('name', '')).strip()[:64]
        if not name:
            continue
        models = [str(m).strip() for m in (b.get('models') or []) if str(m).strip()]
        old = kept.get(name) or {}
        bodies = {m: old[m] for m in models if old.get(m)}
        values.append(
            f"({q(name)}, {qjson(models)}, {qjson(bodies)}, {(i + 1) * 10})"
        )
    if values:
        cur.execute(
            f"INSERT INTO {schema()}.brands (name, models, model_bodies, sort_order) "
            f"VALUES {', '.join(values)} "
            f"ON CONFLICT (name) DO UPDATE SET models = EXCLUDED.models, "
            f"model_bodies = EXCLUDED.model_bodies, sort_order = EXCLUDED.sort_order"
        )
    conn.commit()
    cur.close()
    return len(values)


def make_slug(name: str, taken: set) -> str:
    """Уникальный SEO-адрес товара. При совпадении добавляем номер."""
    base = slugify(name) or 'tovar'
    slug = base
    n = 2
    while slug in taken:
        suffix = f'-{n}'
        slug = base[: SLUG_LIMIT - len(suffix)].rstrip('-') + suffix
        n += 1
    taken.add(slug)
    return slug


def import_categories(conn, items: list) -> int:
    """Сохраняет лист «Категории»: порядок и набор характеристик."""
    if not items:
        return 0
    cur = conn.cursor()
    # Одним запросом на все категории: по строке на каждую было слишком медленно
    values = []
    for i, item in enumerate(items):
        name = str(item.get('name', '')).strip()[:128]
        if not name:
            continue
        order = int(item.get('sortOrder') or 0) or (i + 1) * 10
        fields = [
            str(f).strip()[:64] for f in (item.get('specFields') or []) if str(f).strip()
        ]
        slug = 'cat-' + uuid.uuid4().hex[:8]
        values.append(f"({q(slug)}, {q(name)}, {order}, {qjson(fields)})")
    saved = len(values)
    if values:
        cur.execute(
            f"INSERT INTO {schema()}.categories (slug, name, sort_order, spec_fields) "
            f"VALUES {', '.join(values)} "
            f"ON CONFLICT (name) DO UPDATE SET sort_order = EXCLUDED.sort_order, "
            f"spec_fields = EXCLUDED.spec_fields, is_active = TRUE"
        )
    conn.commit()
    cur.close()
    return saved


def import_rows(conn, in_products: list, in_brands: list, mode: str) -> dict:
    cur = conn.cursor()
    created = 0
    updated = 0

    skipped: list = []
    statements: list = []

    # Забираем все коды товаров разом: искать каждый отдельным запросом
    # слишком долго — на большом каталоге загрузка не успевала завершиться
    cur.execute(f"SELECT slug, id FROM {schema()}.products")
    existing = {row[0]: row[1] for row in cur.fetchall()}
    taken = set(existing)

    for i, p in enumerate(in_products):
        name = str(p.get('name', '')).strip()[:255]
        if not name:
            continue
        slug = str(p.get('slug', '')).strip()[:64] or make_slug(name, taken)
        sku = str(p.get('sku', '')).strip()[:64] or slug.upper()[:64]
        fields = {
            'slug': q(slug),
            'sku': q(sku),
            'name': q(name),
            'category': q(str(p.get('category', '')).strip()[:64] or 'Другое'),
            'price': qint(p.get('price'), 0),
            'old_price': qint(p.get('oldPrice')),
            'pro_price': qint(p.get('proPrice')),
            'ozon_url': q(str(p.get('ozonUrl', ''))),
            'wb_url': q(str(p.get('wbUrl', ''))),
            'install': q(str(p.get('install', ''))[:64]),
            'warranty': q(str(p.get('warranty', ''))[:64]),
            'year_from': qint(p.get('yearFrom'), 2010),
            'year_to': qint(p.get('yearTo'), 2026),
            'badge': q((str(p.get('badge'))[:32] if p.get('badge') else None)),
            'images': qjson(p.get('images') or []),
            'video_url': q(str(p.get('videoUrl') or '')[:500]),
            'description': qjson(p.get('description') or []),
            'specs': qjson(p.get('specs') or []),
            'kit': qjson(p.get('kit') or []),
            'notes': qjson(p.get('notes') or []),
            'extra': qjson(p.get('extra') or []),
            'extra_title': q(str(p.get('extraTitle') or '')[:160]),
            'fits': qjson(p.get('fits') or {}),
            # Пусто — берётся умолчание категории
            'fit_mode': q(p.get('fitMode') if p.get('fitMode') in FIT_MODES else ''),
            'sort_order': qint(p.get('sortOrder'), (i + 1) * 10),
            'popularity': qint(p.get('popularity'), 0),
            'stock_qty': qint(p.get('stock'), 0),
            'stock_note': q(str(p.get('stockNote') or DEFAULT_STOCK_NOTE)[:255]),
            'is_active': 'TRUE' if p.get('isActive', True) else 'FALSE',
        }

        # Столбца нет в файле — у существующего товара оставляем прежнее значение
        full_fields = dict(fields)
        present = p.get('_columns')
        if present:
            keep = {
                'slug': 'slug', 'sku': 'sku', 'name': 'name', 'category': 'category',
                'price': 'price', 'oldPrice': 'old_price', 'proPrice': 'pro_price',
                'ozonUrl': 'ozon_url', 'wbUrl': 'wb_url', 'install': 'install',
                'warranty': 'warranty', 'yearFrom': 'year_from', 'yearTo': 'year_to',
                'badge': 'badge', 'images': 'images', 'description': 'description',
                'specs': 'specs', 'kit': 'kit', 'fits': 'fits',
                'fitMode': 'fit_mode',
                'sortOrder': 'sort_order', 'popularity': 'popularity',
                'stock': 'stock_qty', 'stockNote': 'stock_note',
                'isActive': 'is_active',
            }
            allowed = {keep[k] for k in present if k in keep}
            allowed |= {'slug', 'sku', 'name', 'category'}
            fields = {k: v for k, v in fields.items() if k in allowed}

        found_id = existing.get(slug)
        if found_id:
            if mode == 'skip':
                continue
            sets = ', '.join(f"{k} = {v}" for k, v in fields.items())
            statements.append(
                (i, name, f"UPDATE {schema()}.products SET {sets}, updated_at = NOW() "
                          f"WHERE id = {found_id}")
            )
            updated += 1
        else:
            existing[slug] = -1  # код занят: следующая такая же строка станет обновлением
            statements.append(
                (i, name, f"INSERT INTO {schema()}.products "
                          f"({', '.join(full_fields.keys())}) "
                          f"VALUES ({', '.join(full_fields.values())})")
            )
            created += 1

    # Отправляем все строки одним запросом — так загрузка каталога укладывается
    # в отведённое время. Если где-то ошибка, разбираем построчно, чтобы
    # назвать проблемную строку и сохранить всё остальное
    if statements:
        try:
            cur.execute('; '.join(sql for _i, _n, sql in statements))
        except Exception:
            conn.rollback()
            cur = conn.cursor()
            created = 0
            updated = 0
            for i, name, sql in statements:
                try:
                    cur.execute(sql)
                    if sql.startswith('UPDATE'):
                        updated += 1
                    else:
                        created += 1
                except Exception as exc:
                    conn.rollback()
                    cur = conn.cursor()
                    skipped.append(
                        f'строка {i + 2} ({name[:40]}): {str(exc).strip()[:90]}'
                    )

    # Все марки одним запросом — так загрузка укладывается в отведённое время
    brand_values = []
    for i, b in enumerate(in_brands):
        bname = str(b.get('name', '')).strip()[:64]
        if not bname:
            continue
        models = [str(m).strip() for m in (b.get('models') or []) if str(m).strip()]
        brand_values.append(f"({q(bname)}, {qjson(models)}, {(i + 1) * 10})")
    if brand_values:
        cur.execute(
            f"INSERT INTO {schema()}.brands (name, models, sort_order) "
            f"VALUES {', '.join(brand_values)} "
            f"ON CONFLICT (name) DO UPDATE SET models = EXCLUDED.models"
        )

    # Категории из загруженных товаров должны появиться в справочнике
    cur.execute(
        f"INSERT INTO {schema()}.categories (slug, name, sort_order) "
        f"SELECT 'cat-' || substr(md5(random()::text), 1, 8), src.category, 999 "
        f"FROM (SELECT DISTINCT category FROM {schema()}.products "
        f"WHERE category IS NOT NULL AND category <> '') AS src "
        f"WHERE NOT EXISTS (SELECT 1 FROM {schema()}.categories c WHERE c.name = src.category)"
    )

    conn.commit()
    cur.close()
    result = {'created': created, 'updated': updated}
    if skipped:
        result['skipped'] = len(skipped)
        result['problems'] = skipped[:10]
    return result


def handler(event: dict, context) -> dict:
    """Админка каталога: вход по паролю, список, создание, изменение и удаление товаров."""
    method = event.get('httpMethod', 'GET')
    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': {**CORS, 'Access-Control-Max-Age': '86400'}, 'body': ''}

    params = event.get('queryStringParameters') or {}
    action = params.get('action', '')
    headers = event.get('headers') or {}
    token = headers.get('X-Auth-Token') or headers.get('x-auth-token') or ''
    # Большой список товаров помещается в ответ только сжатым
    _GZIP['on'] = 'gzip' in str(
        headers.get('Accept-Encoding') or headers.get('accept-encoding') or ''
    ).lower()
    raw_body = event.get('body') or '{}'
    try:
        body = json.loads(raw_body) if raw_body else {}
    except json.JSONDecodeError:
        body = {}

    conn = db()
    try:
        if action == 'login':
            password = str(body.get('password', ''))
            real = os.environ.get('ADMIN_PASSWORD', '')
            if not real:
                return resp(500, {'error': 'Пароль администратора не настроен'})
            if hashlib.sha256(password.encode()).hexdigest() != hashlib.sha256(real.encode()).hexdigest():
                return resp(401, {'error': 'Неверный пароль'})
            new_token = secrets.token_hex(32)
            expires = (datetime.utcnow() + timedelta(days=SESSION_DAYS)).strftime('%Y-%m-%d %H:%M:%S')
            cur = conn.cursor()
            cur.execute(
                f"INSERT INTO {schema()}.admin_sessions (token, expires_at) VALUES ({q(new_token)}, {q(expires)})"
            )
            cur.execute(f"DELETE FROM {schema()}.admin_sessions WHERE expires_at < NOW()")
            conn.commit()
            cur.close()
            return resp(200, {'token': new_token})

        if not check_token(conn, token):
            return resp(401, {'error': 'Нужен вход'})

        if action == 'check':
            return resp(200, {'ok': True})

        if action == 'logout':
            cur = conn.cursor()
            cur.execute(f"DELETE FROM {schema()}.admin_sessions WHERE token = {q(token)}")
            conn.commit()
            cur.close()
            return resp(200, {'ok': True})

        if action == 'upload':
            data_url = body.get('image', '')
            if not data_url.startswith('data:'):
                return resp(400, {'error': 'Некорректный файл'})
            return resp(200, {'url': upload_image(data_url)})

        # Видео товара грузится не одним запросом, как фото, а по частям.
        #
        # Один вызов функции принимает не больше ~3,5 МБ тела запроса — это
        # предел самой платформы, обойти его нельзя. Ролик с телефона легко
        # весит 50-150 МБ, поэтому браузер режет файл на куски по 1 МБ и
        # шлёт их один за другим.
        #
        # Штатная многочастевая загрузка S3 (CreateMultipartUpload) в этом
        # хранилище не работает — проверено, отвечает 405. Поэтому каждый
        # кусок кладём отдельным маленьким файлом, а когда пришёл последний,
        # скачиваем все куски по порядку и одним put_object собираем
        # итоговое видео. Временные файлы после сборки удаляем.

        def video_part_key(upload_id: str, index: int) -> str:
            return f"catalog/video-tmp/{upload_id}/{index:05d}"

        if action == 'video-init':
            filename = str(body.get('filename', ''))
            size = int(body.get('size') or 0)
            if size <= 0 or size > MAX_VIDEO_BYTES:
                return resp(400, {'error': 'Видео весит слишком много — до 180 МБ'})
            ext = _video_ext(filename)
            key = f"catalog/video/{uuid.uuid4().hex}.{ext}"
            upload_id = uuid.uuid4().hex
            return resp(200, {'uploadId': upload_id, 'key': key})

        if action == 'video-part':
            upload_id = str(body.get('uploadId', ''))
            index = int(body.get('index') if body.get('index') is not None else -1)
            chunk_b64 = str(body.get('chunk', ''))
            if not upload_id or index < 0 or not chunk_b64:
                return resp(400, {'error': 'Неполные данные куска'})
            try:
                raw = base64.b64decode(chunk_b64)
            except Exception:
                return resp(400, {'error': 'Кусок файла повреждён'})
            _s3().put_object(Bucket='files', Key=video_part_key(upload_id, index), Body=raw)
            return resp(200, {'ok': True})

        if action == 'video-complete':
            key = str(body.get('key', ''))
            upload_id = str(body.get('uploadId', ''))
            total_parts = int(body.get('totalParts') or 0)
            filename = str(body.get('filename', ''))
            if not (key and upload_id and total_parts):
                return resp(400, {'error': 'Неполные данные загрузки'})

            s3 = _s3()

            def fetch_part(i: int) -> bytes:
                return s3.get_object(Bucket='files', Key=video_part_key(upload_id, i))['Body'].read()

            # Части качаем параллельно: это сеть, а не процессор, и по
            # очереди сотня мелких файлов не успела бы собраться в срок
            # жизни функции. Порядок восстанавливаем после — потоки
            # завершаются не по очереди.
            from concurrent.futures import ThreadPoolExecutor

            try:
                with ThreadPoolExecutor(max_workers=16) as pool:
                    chunks = list(pool.map(fetch_part, range(total_parts)))
            except Exception:
                return resp(400, {'error': 'Часть видео не найдена — загрузите заново'})

            data = b''.join(chunks)
            if len(data) > MAX_VIDEO_BYTES:
                return resp(400, {'error': 'Видео весит слишком много — до 180 МБ'})

            ext = _video_ext(filename) if filename else key.rsplit('.', 1)[-1]
            s3.put_object(
                Bucket='files',
                Key=key,
                Body=data,
                ContentType=VIDEO_CONTENT_TYPE.get(ext, 'video/mp4'),
                CacheControl='public, max-age=31536000, immutable',
            )

            def cleanup_part(i: int):
                try:
                    s3.delete_object(Bucket='files', Key=video_part_key(upload_id, i))
                except Exception:
                    pass

            # Удаление временных кусков не влияет на результат — если не
            # уложимся в оставшееся время, лишний мусор в bucket не страшен
            try:
                with ThreadPoolExecutor(max_workers=16) as pool:
                    list(pool.map(cleanup_part, range(total_parts)))
            except Exception:
                pass

            url = f"{CDN_PREFIX}{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"
            return resp(200, {'url': url})

        if action == 'video-abort':
            upload_id = str(body.get('uploadId', ''))
            total_parts = int(body.get('totalParts') or 0)
            if upload_id and total_parts:
                s3 = _s3()
                for i in range(total_parts):
                    try:
                        s3.delete_object(Bucket='files', Key=video_part_key(upload_id, i))
                    except Exception:
                        pass
            return resp(200, {'ok': True})

        if action == 'storage':
            """Ревизия файлового хранилища: что лежит, сколько весит и на
            что сайт больше не ссылается.

            Обход идёт порциями. Хранилище отдаёт список файлов страницами
            по тысяче, а функция живёт считаные секунды — за один вызов всё
            не обойти, поэтому фронт вызывает 'scan' в цикле, пока не
            придёт done. Промежуточный список копится в storage_files.

            Только чтение: этот раздел ничего не удаляет.
            """
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            def folder_of(key: str) -> str:
                """Папка файла — то, что до последней косой черты."""
                return key.rsplit('/', 1)[0] if '/' in key else '(корень)'

            if method == 'POST' and body.get('mode') == 'scan':
                started = time.time()
                s3 = _s3()

                if body.get('restart'):
                    cur.execute(f"TRUNCATE {schema()}.storage_files")
                    cur.execute(f"DELETE FROM {schema()}.storage_scan")
                    cur.execute(
                        f"INSERT INTO {schema()}.storage_scan (id, cursor_key, done) "
                        f"VALUES (1, NULL, FALSE)"
                    )
                    conn.commit()

                cur.execute(
                    f"SELECT cursor_key, done FROM {schema()}.storage_scan WHERE id = 1"
                )
                state = cur.fetchone()
                if not state:
                    cur.execute(
                        f"INSERT INTO {schema()}.storage_scan (id, cursor_key, done) "
                        f"VALUES (1, NULL, FALSE)"
                    )
                    conn.commit()
                    state = {'cursor_key': None, 'done': False}

                token = state['cursor_key']
                added = 0
                done = False

                # Успеваем сколько успеем: 3 секунды с запасом до таймаута
                while time.time() - started < 3.0:
                    kw = {'Bucket': 'files', 'MaxKeys': 1000}
                    if token:
                        kw['StartAfter'] = token
                    page = s3.list_objects_v2(**kw)
                    items = page.get('Contents') or []
                    if not items:
                        done = True
                        # Хранилище не отдаёт список файлов: обзор каталога
                        # для наших ключей закрыт, доступ есть только к
                        # конкретному файлу по имени. Отличаем это от честно
                        # пустого хранилища, чтобы не показать «0 файлов»
                        # там, где их тысячи.
                        if not token:
                            cur.close()
                            return resp(200, {
                                'done': True,
                                'scanned': 0,
                                'added': 0,
                                'unsupported': True,
                            })
                        break

                    values = []
                    for it in items:
                        k = it['Key']
                        size = int(it.get('Size') or 0)
                        mod = it.get('LastModified')
                        mod_sql = q(mod.isoformat()) if mod else 'NULL'
                        values.append(f"({q(k)}, {size}, {mod_sql})")
                    cur.execute(
                        f"INSERT INTO {schema()}.storage_files (key, size_bytes, modified_at) "
                        f"VALUES {','.join(values)} "
                        f"ON CONFLICT (key) DO UPDATE SET "
                        f"size_bytes = EXCLUDED.size_bytes, "
                        f"modified_at = EXCLUDED.modified_at"
                    )
                    added += len(items)
                    token = items[-1]['Key']

                    if not page.get('IsTruncated'):
                        done = True
                        break

                cur.execute(
                    f"UPDATE {schema()}.storage_scan SET cursor_key = {q(token) if token else 'NULL'}, "
                    f"done = {'TRUE' if done else 'FALSE'}, "
                    f"finished_at = {'NOW()' if done else 'NULL'} WHERE id = 1"
                )
                conn.commit()

                if done:
                    # Полный список знаем только сейчас — размечаем, что нужно
                    # сайту. Ссылки лежат в разных таблицах и внутри JSON,
                    # поэтому собираем их одним запросом и сверяем по имени файла.
                    cdn_pref = f"{CDN_PREFIX}{os.environ.get('AWS_ACCESS_KEY_ID', '')}/bucket/"
                    used_sql = (
                        f"SELECT jsonb_array_elements_text(images) AS u "
                        f"FROM {schema()}.products WHERE images IS NOT NULL "
                        f"UNION ALL SELECT video_url FROM {schema()}.products "
                        f"UNION ALL SELECT scheme_url FROM {schema()}.products "
                        f"UNION ALL SELECT b->>'image' FROM {schema()}.products, "
                        f"jsonb_array_elements(COALESCE(notes, '[]'::jsonb)) b "
                        f"UNION ALL SELECT b->>'image' FROM {schema()}.products, "
                        f"jsonb_array_elements(COALESCE(extra, '[]'::jsonb)) b "
                        f"UNION ALL SELECT b->>'video' FROM {schema()}.products, "
                        f"jsonb_array_elements(COALESCE(extra, '[]'::jsonb)) b "
                        f"UNION ALL SELECT b->>'video' FROM {schema()}.products, "
                        f"jsonb_array_elements(COALESCE(notes, '[]'::jsonb)) b "
                        f"UNION ALL SELECT cover FROM {schema()}.guides "
                        f"UNION ALL SELECT video_url FROM {schema()}.guides "
                        f"UNION ALL SELECT b->>'image' FROM {schema()}.guides, "
                        f"jsonb_array_elements(COALESCE(blocks, '[]'::jsonb)) b "
                        f"UNION ALL SELECT b->>'video' FROM {schema()}.guides, "
                        f"jsonb_array_elements(COALESCE(blocks, '[]'::jsonb)) b "
                        f"UNION ALL SELECT image FROM {schema()}.categories"
                    )
                    cur.execute(
                        f"UPDATE {schema()}.storage_files f SET used = EXISTS ("
                        f"SELECT 1 FROM ({used_sql}) t "
                        f"WHERE t.u IS NOT NULL AND t.u <> '' "
                        f"AND split_part(t.u, '?', 1) = {q(cdn_pref)} || f.key)"
                    )
                    conn.commit()

                cur.execute(f"SELECT COUNT(*) AS n FROM {schema()}.storage_files")
                total = int(cur.fetchone()['n'] or 0)
                cur.close()
                return resp(200, {'done': done, 'scanned': total, 'added': added})

            # GET — показать последний отчёт
            cur.execute(
                f"SELECT cursor_key, done, started_at, finished_at "
                f"FROM {schema()}.storage_scan WHERE id = 1"
            )
            scan = cur.fetchone()

            cur.execute(
                f"SELECT COUNT(*) AS files, COALESCE(SUM(size_bytes), 0) AS bytes, "
                f"COUNT(*) FILTER (WHERE used) AS used_files, "
                f"COALESCE(SUM(size_bytes) FILTER (WHERE used), 0) AS used_bytes, "
                f"COUNT(*) FILTER (WHERE used IS FALSE) AS free_files, "
                f"COALESCE(SUM(size_bytes) FILTER (WHERE used IS FALSE), 0) AS free_bytes "
                f"FROM {schema()}.storage_files"
            )
            totals = cur.fetchone() or {}

            # Группируем максимум по двум уровням: обрывки видео лежат
            # каждый в своей подпапке, и по полному пути отчёт превратился
            # бы в сотни одинаковых строк вместо одной понятной
            folder_expr = (
                "CASE WHEN POSITION('/' IN key) = 0 THEN '(корень)' "
                "WHEN POSITION('/' IN SUBSTRING(key FROM POSITION('/' IN key) + 1)) = 0 "
                "THEN SPLIT_PART(key, '/', 1) "
                "ELSE SPLIT_PART(key, '/', 1) || '/' || SPLIT_PART(key, '/', 2) END"
            )
            cur.execute(
                f"SELECT {folder_expr} AS folder, "
                f"COUNT(*) AS files, COALESCE(SUM(size_bytes), 0) AS bytes, "
                f"COUNT(*) FILTER (WHERE used) AS used_files, "
                f"COUNT(*) FILTER (WHERE used IS FALSE) AS free_files, "
                f"COALESCE(SUM(size_bytes) FILTER (WHERE used IS FALSE), 0) AS free_bytes, "
                f"MIN(modified_at) AS oldest, MAX(modified_at) AS newest "
                f"FROM {schema()}.storage_files GROUP BY 1 ORDER BY bytes DESC"
            )
            folders = [
                {
                    'folder': r['folder'],
                    'files': int(r['files']),
                    'bytes': int(r['bytes']),
                    'usedFiles': int(r['used_files'] or 0),
                    'freeFiles': int(r['free_files'] or 0),
                    'freeBytes': int(r['free_bytes'] or 0),
                    'oldest': r['oldest'].isoformat() if r['oldest'] else None,
                    'newest': r['newest'].isoformat() if r['newest'] else None,
                }
                for r in cur.fetchall()
            ]

            # Картина по каталогу — она доступна всегда, даже когда обзор
            # хранилища закрыт: считаем ссылки, которыми пользуется сайт.
            cdn_pref_now = f"{CDN_PREFIX}{os.environ.get('AWS_ACCESS_KEY_ID', '')}/bucket/"
            refs_sql = (
                f"SELECT jsonb_array_elements_text(images) AS u "
                f"FROM {schema()}.products WHERE images IS NOT NULL "
                f"UNION ALL SELECT video_url FROM {schema()}.products "
                f"UNION ALL SELECT scheme_url FROM {schema()}.products "
                f"UNION ALL SELECT b->>'image' FROM {schema()}.products, "
                f"jsonb_array_elements(COALESCE(notes, '[]'::jsonb)) b "
                f"UNION ALL SELECT b->>'image' FROM {schema()}.products, "
                f"jsonb_array_elements(COALESCE(extra, '[]'::jsonb)) b "
                f"UNION ALL SELECT b->>'video' FROM {schema()}.products, "
                f"jsonb_array_elements(COALESCE(extra, '[]'::jsonb)) b "
                f"UNION ALL SELECT cover FROM {schema()}.guides "
                f"UNION ALL SELECT video_url FROM {schema()}.guides "
                f"UNION ALL SELECT b->>'image' FROM {schema()}.guides, "
                f"jsonb_array_elements(COALESCE(blocks, '[]'::jsonb)) b "
                f"UNION ALL SELECT b->>'video' FROM {schema()}.guides, "
                f"jsonb_array_elements(COALESCE(blocks, '[]'::jsonb)) b "
                f"UNION ALL SELECT image FROM {schema()}.categories"
            )
            cur.execute(
                f"SELECT COUNT(*) AS refs, COUNT(DISTINCT u) AS files, "
                f"COUNT(DISTINCT u) FILTER (WHERE u LIKE {q(cdn_pref_now + '%')}) AS own, "
                f"COUNT(DISTINCT u) FILTER (WHERE u NOT LIKE {q(cdn_pref_now + '%')}) AS external, "
                f"COUNT(DISTINCT u) FILTER (WHERE u LIKE {q(cdn_pref_now + 'catalog/video/%')}) AS videos "
                f"FROM ({refs_sql}) t WHERE u IS NOT NULL AND u <> ''"
            )
            cat = cur.fetchone() or {}

            cur.close()
            return resp(200, {
                'catalog': {
                    'refs': int(cat.get('refs') or 0),
                    'files': int(cat.get('files') or 0),
                    'own': int(cat.get('own') or 0),
                    'external': int(cat.get('external') or 0),
                    'videos': int(cat.get('videos') or 0),
                },
                'scan': {
                    'done': bool(scan['done']) if scan else False,
                    'startedAt': scan['started_at'].isoformat() if scan and scan['started_at'] else None,
                    'finishedAt': scan['finished_at'].isoformat() if scan and scan['finished_at'] else None,
                } if scan else None,
                'totals': {
                    'files': int(totals.get('files') or 0),
                    'bytes': int(totals.get('bytes') or 0),
                    'usedFiles': int(totals.get('used_files') or 0),
                    'usedBytes': int(totals.get('used_bytes') or 0),
                    'freeFiles': int(totals.get('free_files') or 0),
                    'freeBytes': int(totals.get('free_bytes') or 0),
                },
                'folders': folders,
            })

        if action == 'external-images':
            """Переносит к нам картинки, которые в каталоге указаны ссылками
            на чужие сайты. За один вызов — сколько успеем до конца отведённого
            функции времени: скачивание идёт через интернет и непредсказуемо."""
            started = time.time()
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            if method == 'DELETE':
                # Сбросить список неудачных — чтобы попробовать заново
                cur.execute(f"DELETE FROM {schema()}.failed_images")
                conn.commit()
                cur.close()
                return resp(200, {'ok': True})

            cur.execute(f"SELECT url, reason, product_slug FROM {schema()}.failed_images")
            failed_rows = cur.fetchall()
            skip = {r['url'] for r in failed_rows}

            # Берём только товары, где реально есть чужие ссылки.
            # Отбор делает база: искать «http» по всему полю нельзя —
            # наши собственные адреса тоже начинаются с http, и в выборку
            # попадали уже перенесённые товары, вытесняя те, что ждут очереди.
            #
            # Ссылки из чёрного списка тоже не считаются: у первых сорока
            # товаров каталога все картинки оказались битыми, выборка
            # упиралась в них и возвращала «нечего переносить», хотя дальше
            # ждали тысячи рабочих. Поэтому условие требует хотя бы одну
            # ссылку, которую ещё имеет смысл качать.
            not_failed = (
                f"NOT EXISTS (SELECT 1 FROM {schema()}.failed_images f "
                f"WHERE f.url = u)"
            )
            has_external = (
                f"EXISTS (SELECT 1 FROM jsonb_array_elements_text(images) AS u "
                f"WHERE u LIKE 'http%' AND u NOT LIKE {q(CDN_PREFIX + '%')} "
                f"AND {not_failed})"
            )
            tail = '' if method == 'GET' else ' LIMIT 40'
            cur.execute(
                f"SELECT id, slug, images FROM {schema()}.products "
                f"WHERE {has_external} ORDER BY id{tail}"
            )
            rows = cur.fetchall()

            def pending_of(r):
                return [
                    u for u in (r['images'] or [])
                    if _is_external_image(u) and u not in skip
                ]

            if method == 'GET':
                left = sum(len(pending_of(r)) for r in rows)
            else:
                # При переносе в выборке только часть товаров, поэтому
                # общий остаток берём отдельным быстрым запросом
                # Считаем только то, что реально предстоит скачать:
                # отказы уже исключены условием выборки
                cur.execute(
                    f"SELECT COALESCE(SUM(("
                    f"SELECT COUNT(*) FROM jsonb_array_elements_text(images) AS u "
                    f"WHERE u LIKE 'http%' AND u NOT LIKE {q(CDN_PREFIX + '%')} "
                    f"AND {not_failed}"
                    f")), 0) AS n FROM {schema()}.products "
                    f"WHERE {has_external}"
                )
                left = max(int(cur.fetchone()['n'] or 0), 0)

            failed = [
                {
                    'url': r['url'],
                    'reason': r['reason'],
                    'product': r['product_slug'],
                }
                for r in failed_rows
            ]

            if method == 'GET':
                cur.close()
                return resp(200, {'left': left, 'failed': failed})

            # POST — переносим столько картинок, сколько успеем.
            # Раньше за вызов бралась ровно одна, и почти всё время уходило
            # на разогрев функции; теперь работаем до конца отведённого срока.
            saved = 0
            problem = None
            done_urls = set()

            # Собираем очередь ссылок: качать будем пачкой, а не по одной.
            # Сеть — самая долгая часть, и раньше за вызов успевала пройти
            # одна картинка. Параллельно за то же время проходит десяток.
            queue = []
            for r in rows:
                for idx, u in enumerate(r['images'] or []):
                    if _is_external_image(u) and u not in skip:
                        queue.append((r, idx, u))

            spare = FUNCTION_BUDGET - RESERVE - (time.time() - started)
            batch = queue[:BATCH_IMAGES]

            if batch and spare > 0.4:
                from concurrent.futures import ThreadPoolExecutor

                # Половину остатка отдаём сети, половину бережём на сжатие
                # и отправку в хранилище — иначе скачаем и не успеем сохранить
                wait = max(min(spare * 0.5, 1.0), 0.4)
                with ThreadPoolExecutor(max_workers=len(batch)) as ex:
                    got = list(
                        ex.map(lambda it: download_external(it[2], wait), batch)
                    )

                # Скачанное складываем к нам и правим ссылки в товарах
                changed_rows = {}
                for (r, idx, u), (raw, ext, reason) in zip(batch, got):
                    # Время вышло — необработанное не помечаем разобранным,
                    # чтобы вернуться к нему на следующем заходе
                    if FUNCTION_BUDGET - RESERVE - (time.time() - started) < 0.2:
                        break
                    done_urls.add(u)
                    own = None
                    if raw:
                        own, reason = store_image(raw, ext)

                    if own:
                        urls = changed_rows.setdefault(r['id'], list(r['images']))
                        urls[idx] = own
                        saved += 1
                    elif reason == 'сайт отвечает слишком долго':
                        # Не вина ссылки — сайт не успел ответить.
                        # В список отказов не заносим: попробуем позже.
                        problem = {
                            'url': u, 'reason': reason, 'product': r['slug'],
                        }
                    else:
                        # Запоминаем отказ, чтобы не биться в эту ссылку снова
                        cur.execute(
                            f"INSERT INTO {schema()}.failed_images "
                            f"(url, reason, product_slug) VALUES "
                            f"({q(u)}, {q(reason or 'неизвестная ошибка')}, {q(r['slug'])}) "
                            f"ON CONFLICT (url) DO UPDATE SET tries = failed_images.tries + 1, "
                            f"reason = EXCLUDED.reason"
                        )
                        skip.add(u)
                        problem = {
                            'url': u, 'reason': reason, 'product': r['slug'],
                        }

                for pid, urls in changed_rows.items():
                    cur.execute(
                        f"UPDATE {schema()}.products SET images = {qjson(urls)}, "
                        f"updated_at = NOW() WHERE id = {pid}"
                    )
                conn.commit()

            cur.close()
            # Сколько ещё осталось — считаем от того, что реально разобрали
            handled = len(done_urls)
            return resp(200, {
                'saved': saved,
                'left': max(left - handled, 0),
                'handled': handled,
                'problem': problem,
            })

        if action == 'optimize-images':
            """Пережимает фото товаров в WebP порциями, чтобы уложиться в таймаут."""
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(f"SELECT url FROM {schema()}.optimize_skip")
            skip_opt = {r['url'] for r in cur.fetchall()}

            # Только наши файлы: у чужих ссылок расширение тоже .jpg,
            # но пережать их нельзя — сначала перенос в своё хранилище
            cur.execute(
                f"SELECT id, images FROM {schema()}.products "
                f"WHERE EXISTS (SELECT 1 FROM jsonb_array_elements_text(images) AS u "
                f"WHERE u LIKE {q(CDN_PREFIX + '%')} AND lower(split_part(u, '?', 1)) "
                f"~ '\\.(png|jpe?g)$' AND NOT EXISTS (SELECT 1 FROM {schema()}.optimize_skip s "
                f"WHERE s.url = u)) ORDER BY id"
            )
            rows = cur.fetchall()

            def heavy(u: str) -> bool:
                """Фото, которое мы можем пережать.

                Чужие ссылки сюда не годятся: пережимаем мы только то, что
                лежит в нашем хранилище. Раньше их тоже считали, и счётчик
                вечно висел на «не оптимизировано» — обработать эти файлы
                было нельзя, пока они не переехали к нам.
                """
                if not _is_own_cdn(u) or u in skip_opt:
                    return False
                return u.lower().split('?')[0].endswith(('.png', '.jpg', '.jpeg'))

            left = sum(len([u for u in (r['images'] or []) if heavy(u)]) for r in rows)

            # Чужие фото ждут переноса, а не сжатия — показываем отдельно,
            # чтобы было видно: работа есть, но сначала другая
            cur.execute(
                f"SELECT COALESCE(SUM(("
                f"SELECT COUNT(*) FROM jsonb_array_elements_text(images) AS u "
                f"WHERE u LIKE 'http%' AND u NOT LIKE {q(CDN_PREFIX + '%')}"
                f")), 0) AS n FROM {schema()}.products"
            )
            external = int(cur.fetchone()['n'] or 0)

            if method == 'GET':
                cur.close()
                return resp(200, {'left': left, 'external': external})

            # За один вызов пережимаем ровно одно фото — лимит функции 2 секунды
            saved = 0
            for r in rows:
                urls = r['images'] or []
                idx = next((i for i, u in enumerate(urls) if heavy(u)), None)
                if idx is None:
                    continue
                new_url = reoptimize_url(urls[idx])
                if new_url == urls[idx]:
                    # Сжатие не дало выигрыша — файл и так лёгкий. Помечаем,
                    # иначе счётчик застрянет на нём навсегда: каждый вызов
                    # брал бы то же фото и снова упирался в тот же отказ
                    cur.execute(
                        f"INSERT INTO {schema()}.optimize_skip (url, reason) "
                        f"VALUES ({q(urls[idx])}, {q('сжатие не уменьшает файл')}) "
                        f"ON CONFLICT (url) DO NOTHING"
                    )
                    conn.commit()
                    continue
                urls[idx] = new_url
                cur.execute(
                    f"UPDATE {schema()}.products SET images = {qjson(urls)} "
                    f"WHERE id = {r['id']}"
                )
                conn.commit()
                saved = 1
                break

            cur.close()
            return resp(200, {
                'done': saved,
                'saved': saved,
                'left': max(left - saved, 0),
                'external': external,
            })

        if action == 'bulk' and method == 'POST':
            ids = [int(i) for i in (body.get('ids') or []) if str(i).isdigit()]
            if not ids:
                return resp(400, {'error': 'Не выбрано ни одного товара'})
            id_list = ','.join(str(i) for i in ids)
            op = str(body.get('op', ''))
            cur = conn.cursor()

            if op == 'category':
                category = str(body.get('category', '')).strip()[:64]
                if not category:
                    cur.close()
                    return resp(400, {'error': 'Не указана категория'})
                cur.execute(
                    f"UPDATE {schema()}.products SET category = {q(category)}, "
                    f"updated_at = NOW() WHERE id IN ({id_list})"
                )
            elif op == 'delete':
                cur.execute(f"DELETE FROM {schema()}.product_guides WHERE product_id IN ({id_list})")
                cur.execute(f"DELETE FROM {schema()}.products WHERE id IN ({id_list})")
            elif op == 'frame-wires':
                # Проводки сразу всей группе рамок одного периода: на 9",
                # 10" и 12,3" под одну машину проводка одна и та же, и
                # проставлять её каждой отдельно — тройная работа
                slugs = [
                    str(x).strip()
                    for x in (body.get('frameWires') or [])
                    if str(x).strip()
                ][:20]
                cur.execute(
                    f"UPDATE {schema()}.products SET frame_wires = {qjson(slugs)}, "
                    f"updated_at = NOW() WHERE id IN ({id_list})"
                )
            elif op == 'wire-tech':
                # Разметка проводок пачкой: у похожих позиций набор
                # «камера / усилитель / CAN» совпадает, и щёлкать каждую
                # по отдельности — терять время на 434 товарах
                tech = clean_wire_tech(body.get('wireTech') or {})
                cur.execute(
                    f"UPDATE {schema()}.products SET wire_tech = {qjson(tech)}, "
                    f"updated_at = NOW() WHERE id IN ({id_list})"
                )
            elif op in ('show', 'hide'):
                flag = 'TRUE' if op == 'show' else 'FALSE'
                cur.execute(
                    f"UPDATE {schema()}.products SET is_active = {flag}, "
                    f"updated_at = NOW() WHERE id IN ({id_list})"
                )
            else:
                cur.close()
                return resp(400, {'error': 'Неизвестное действие'})

            conn.commit()
            cur.close()
            return resp(200, {'ok': True, 'affected': len(ids)})

        if action == 'categories':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            if method == 'GET':
                cur.execute(
                    f"SELECT c.name, c.sort_order, c.is_active, c.spec_fields, c.fit_mode, "
                    f"(SELECT COUNT(*) FROM {schema()}.products p WHERE p.category = c.name) AS products "
                    f"FROM {schema()}.categories c WHERE c.is_active "
                    f"ORDER BY c.sort_order, c.name"
                )
                items = [
                    {
                        'name': r['name'],
                        'sortOrder': r['sort_order'],
                        'products': int(r['products']),
                        'specFields': r['spec_fields'] or [],
                        'fitMode': r['fit_mode'] or 'universal',
                    }
                    for r in cur.fetchall()
                ]
                cur.close()
                return resp(200, {'categories': items})

            if method == 'PUT':
                items = body.get('categories') or []
                # Прячем все, затем возвращаем присланные — так удаление не теряет товары
                cur.execute(f"UPDATE {schema()}.categories SET is_active = FALSE")
                for i, item in enumerate(items):
                    name = str(item.get('name', '')).strip()[:128]
                    if not name:
                        continue
                    old = str(item.get('oldName', '')).strip()[:128]
                    order = (i + 1) * 10
                    fields = [
                        str(f).strip()[:64]
                        for f in (item.get('specFields') or [])
                        if str(f).strip()
                    ]
                    mode = item.get('fitMode')
                    mode = q(mode if mode in FIT_MODES else 'universal')
                    if old and old != name:
                        # Переименование — тянем за собой товары
                        cur.execute(
                            f"UPDATE {schema()}.products SET category = {q(name)}, "
                            f"updated_at = NOW() WHERE category = {q(old)}"
                        )
                        cur.execute(
                            f"UPDATE {schema()}.categories SET name = {q(name)}, "
                            f"sort_order = {order}, is_active = TRUE, "
                            f"spec_fields = {qjson(fields)}, fit_mode = {mode} "
                            f"WHERE name = {q(old)}"
                        )
                        continue
                    cur.execute(
                        f"SELECT id FROM {schema()}.categories WHERE name = {q(name)}"
                    )
                    if cur.fetchone():
                        cur.execute(
                            f"UPDATE {schema()}.categories SET sort_order = {order}, "
                            f"is_active = TRUE, spec_fields = {qjson(fields)}, "
                            f"fit_mode = {mode} WHERE name = {q(name)}"
                        )
                    else:
                        slug = 'cat-' + uuid.uuid4().hex[:8]
                        cur.execute(
                            f"INSERT INTO {schema()}.categories "
                            f"(slug, name, sort_order, spec_fields, fit_mode) "
                            f"VALUES ({q(slug)}, {q(name)}, {order}, {qjson(fields)}, {mode})"
                        )
                conn.commit()
                cur.close()
                return resp(200, {'ok': True})

            cur.close()
            return resp(400, {'error': 'Неизвестное действие'})

        if action == 'missing-fit':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            if method == 'GET':
                cur.execute(
                    f"SELECT id, brand, model, year, scenario, contact, hits, "
                    f"created_at, updated_at FROM {schema()}.missing_fit_requests "
                    f"ORDER BY updated_at DESC LIMIT 500"
                )
                items = [
                    {
                        'id': r['id'],
                        'brand': r['brand'],
                        'model': r['model'],
                        'year': r['year'],
                        'scenario': r['scenario'],
                        'contact': r['contact'],
                        'hits': r['hits'],
                        'createdAt': r['created_at'].isoformat() if r['created_at'] else None,
                        'updatedAt': r['updated_at'].isoformat() if r['updated_at'] else None,
                    }
                    for r in cur.fetchall()
                ]
                cur.close()
                return resp(200, {'items': items})

            if method == 'DELETE':
                rid = params.get('id', '')
                if not str(rid).isdigit():
                    cur.close()
                    return resp(400, {'error': 'Не указана запись'})
                cur.execute(
                    f"DELETE FROM {schema()}.missing_fit_requests WHERE id = {int(rid)}"
                )
                conn.commit()
                cur.close()
                return resp(200, {'ok': True})

            cur.close()
            return resp(400, {'error': 'Неизвестное действие'})

        if action == 'dealers':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            if method == 'GET':
                cur.execute(
                    f"SELECT id, phone, name, comment, is_active, last_login, created_at "
                    f"FROM {schema()}.dealers ORDER BY created_at DESC"
                )
                items = [
                    {
                        'id': r['id'],
                        'phone': r['phone'],
                        'name': r['name'],
                        'comment': r['comment'],
                        'isActive': r['is_active'],
                        'lastLogin': r['last_login'].isoformat() if r['last_login'] else None,
                        'createdAt': r['created_at'].isoformat() if r['created_at'] else None,
                    }
                    for r in cur.fetchall()
                ]
                cur.close()
                return resp(200, {'dealers': items})

            if method == 'POST':
                phone = re.sub(r'\D', '', str(body.get('phone', '')))
                if phone.startswith('8'):
                    phone = '7' + phone[1:]
                if len(phone) == 10:
                    phone = '7' + phone
                if len(phone) != 11:
                    cur.close()
                    return resp(400, {'error': 'Введите номер полностью'})

                name = str(body.get('name', ''))[:160]
                comment = str(body.get('comment', ''))[:255]
                cur.execute(
                    f"INSERT INTO {schema()}.dealers (phone, name, comment) "
                    f"VALUES ({q(phone)}, {q(name)}, {q(comment)}) "
                    f"ON CONFLICT (phone) DO UPDATE SET name = EXCLUDED.name, "
                    f"comment = EXCLUDED.comment, is_active = TRUE"
                )
                conn.commit()
                cur.close()
                return resp(200, {'ok': True})

            if method == 'DELETE':
                dealer_id = int(body.get('id') or 0)
                if dealer_id:
                    cur.execute(f"DELETE FROM {schema()}.dealers WHERE id = {dealer_id}")
                    conn.commit()
                cur.close()
                return resp(200, {'ok': True})

            if method == 'PUT':
                dealer_id = int(body.get('id') or 0)
                active = 'TRUE' if body.get('isActive') else 'FALSE'
                if dealer_id:
                    cur.execute(
                        f"UPDATE {schema()}.dealers SET is_active = {active} WHERE id = {dealer_id}"
                    )
                    conn.commit()
                cur.close()
                return resp(200, {'ok': True})

            cur.close()

        if action == 'vehicle-wiring' and method == 'GET':
            # Настройки подбора проводки — для вкладки «Марки»
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(
                f"SELECT id, brand, model, year_from, year_to, mode, wire_slug, "
                f"reason, ask, wheel, bodies FROM {schema()}.vehicle_wiring "
                f"ORDER BY brand, model, year_from"
            )
            rows = [
                {
                    'id': r['id'],
                    'brand': r['brand'],
                    'model': r['model'],
                    'yearFrom': r['year_from'],
                    'yearTo': r['year_to'],
                    'mode': r['mode'],
                    'wireSlug': r['wire_slug'],
                    'reason': r['reason'],
                    'ask': r['ask'] or {},
                    'wheel': r['wheel'] or '',
                    'bodies': r['bodies'] or [],
                }
                for r in cur.fetchall()
            ]
            cur.execute(
                f"SELECT slug, sku, name, price FROM {schema()}.products "
                f"WHERE category = {q(WIRES_CATEGORY)} AND is_active "
                f"ORDER BY name"
            )
            wires = [
                {
                    'slug': w['slug'],
                    'sku': w['sku'] or '',
                    'name': w['name'],
                    'price': w['price'],
                }
                for w in cur.fetchall()
            ]
            cur.close()
            return resp(200, {'rows': rows, 'wires': wires})

        if action == 'vehicle-wiring' and method == 'DELETE':
            rid = qint(body.get('id'))
            if rid == 'NULL':
                return resp(400, {'error': 'Не указана строка'})
            cur = conn.cursor()
            cur.execute(
                f"DELETE FROM {schema()}.vehicle_wiring WHERE id = {rid}"
            )
            conn.commit()
            cur.close()
            return resp(200, {'ok': True})

        if action == 'vehicle-wiring' and method in ('PUT', 'POST'):
            # Одно поколение машины за раз. У модели их может быть несколько:
            # Civic до 2011 — хэтчбек, после — седан, и проводки разные
            brand = str(body.get('brand') or '').strip()
            model = str(body.get('model') or '').strip()
            if not brand or not model:
                return resp(400, {'error': 'Не указана машина'})
            mode = body.get('mode')
            if mode not in WIRE_MODES:
                return resp(400, {'error': 'Неизвестный режим подбора'})
            ask = body.get('ask') or {}
            bodies = [
                b for b in (body.get('bodies') or []) if b in BODY_TYPES
            ]
            fields = (
                f"brand = {q(brand)}, model = {q(model)}, "
                f"year_from = {qint(body.get('yearFrom'), 1990)}, "
                f"year_to = {qint(body.get('yearTo'), 2100)}, "
                f"mode = {q(mode)}, "
                f"wire_slug = {q(str(body.get('wireSlug') or '')[:160])}, "
                f"reason = {q(str(body.get('reason') or '')[:600])}, "
                f"ask = {qjson({k: bool(ask.get(k)) for k in ('amp', 'camera', 'can')})}, "
                f"wheel = {q(body.get('wheel') if body.get('wheel') in WHEEL_SIDES else '')}, "
                f"bodies = {qjson(bodies)}, updated_at = NOW()"
            )
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            rid = qint(body.get('id'))
            if rid != 'NULL':
                cur.execute(
                    f"UPDATE {schema()}.vehicle_wiring SET {fields} "
                    f"WHERE id = {rid} RETURNING id"
                )
            else:
                cur.execute(
                    f"INSERT INTO {schema()}.vehicle_wiring (brand, model) "
                    f"VALUES ({q(brand)}, {q(model)}) RETURNING id"
                )
                new_id = cur.fetchone()['id']
                cur.execute(
                    f"UPDATE {schema()}.vehicle_wiring SET {fields} "
                    f"WHERE id = {new_id} RETURNING id"
                )
            row = cur.fetchone()
            conn.commit()
            cur.close()
            return resp(200, {'ok': True, 'id': row['id'] if row else None})

        if action == 'brands' and method == 'PUT':
            brands = body.get('brands', [])
            cur = conn.cursor()
            cur.execute(f"DELETE FROM {schema()}.brands")
            # Копим строки и вставляем разом: по одному запросу на марку
            # сохранение справочника упиралось в таймаут и падало с ошибкой
            values = []
            for i, b in enumerate(brands):
                name = str(b.get('name', '')).strip()
                if not name:
                    continue
                models = [str(m).strip() for m in b.get('models', []) if str(m).strip()]
                # Типы кузова по моделям: {"Rio": ["sedan", "hatchback"]}.
                # Держим только те модели, что реально есть в списке марки.
                raw_bodies = b.get('modelBodies') or {}
                bodies = {}
                if isinstance(raw_bodies, dict):
                    known = set(models)
                    for model, kinds in raw_bodies.items():
                        model = str(model).strip()
                        if model not in known or not isinstance(kinds, list):
                            continue
                        clean = [
                            str(k).strip()
                            for k in kinds
                            if str(k).strip() in BODY_TYPES
                        ]
                        if clean:
                            bodies[model] = clean
                values.append(
                    f"({q(name)}, {qjson(models)}, {qjson(bodies)}, {(i + 1) * 10})"
                )
            if values:
                cur.execute(
                    f"INSERT INTO {schema()}.brands "
                    f"(name, models, model_bodies, sort_order) "
                    f"VALUES {', '.join(values)}"
                )
            conn.commit()
            cur.close()
            return resp(200, {'ok': True})

        if action == 'brands-export-xlsx':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(
                f"SELECT name, models FROM {schema()}.brands ORDER BY sort_order, id"
            )
            brands = [{'name': b['name'], 'models': b['models']} for b in cur.fetchall()]
            cur.close()
            data = build_brands_xlsx(brands)
            return resp(200, {'file': base64.b64encode(data).decode('ascii')})

        if action == 'brands-import-xlsx':
            raw = str(body.get('file', ''))
            if ',' in raw and raw.startswith('data:'):
                raw = raw.split(',', 1)[1]
            try:
                blob = base64.b64decode(raw)
            except Exception:
                return resp(400, {'error': 'Файл не читается'})
            try:
                parsed = parse_brands_xlsx(blob)
            except Exception:
                return resp(400, {'error': 'Это не таблица Excel или структура изменена'})
            if not parsed:
                return resp(400, {'error': 'В таблице не нашлось ни одной марки'})
            saved = replace_brands(conn, parsed)
            models = sum(len(b['models']) for b in parsed)
            return resp(200, {'ok': True, 'brands': saved, 'models': models})

        if action == 'settings':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            if method == 'GET':
                cur.execute(f"SELECT key, value FROM {schema()}.settings")
                data = {s['key']: s['value'] for s in cur.fetchall()}
                cur.close()
                return resp(200, {'settings': data})
            if method == 'PUT':
                for key, value in (body.get('settings') or {}).items():
                    cur.execute(
                        f"INSERT INTO {schema()}.settings (key, value, updated_at) "
                        f"VALUES ({q(str(key)[:64])}, {qjson(value)}, NOW()) "
                        f"ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()"
                    )
                conn.commit()
                cur.close()
                return resp(200, {'ok': True})
            cur.close()
            return resp(400, {'error': 'Неизвестное действие'})

        if action == 'export':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(f"SELECT * FROM {schema()}.products ORDER BY sort_order, id")
            products = [row_to_product(r) for r in cur.fetchall()]
            cur.execute(f"SELECT name, models FROM {schema()}.brands ORDER BY sort_order, id")
            brands = [{'name': b['name'], 'models': b['models']} for b in cur.fetchall()]
            cur.close()
            for p in products:
                p.pop('id', None)
            return resp(200, {'version': 1, 'products': products, 'brands': brands})

        if action == 'import':
            stats = import_rows(
                conn,
                body.get('products') or [],
                body.get('brands') or [],
                str(body.get('mode', 'merge')),
            )
            return resp(200, {'ok': True, **stats})

        if action == 'export-xlsx':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(f"SELECT * FROM {schema()}.products ORDER BY sort_order, id")
            products = [row_to_product(r) for r in cur.fetchall()]
            cur.execute(f"SELECT name, models FROM {schema()}.brands ORDER BY sort_order, id")
            brands = [{'name': b['name'], 'models': b['models']} for b in cur.fetchall()]
            cur.execute(
                f"SELECT name, sort_order, spec_fields FROM {schema()}.categories "
                f"WHERE is_active ORDER BY sort_order, name"
            )
            categories = [
                {
                    'name': c['name'],
                    'sortOrder': c['sort_order'],
                    'specFields': c['spec_fields'] or [],
                }
                for c in cur.fetchall()
            ]
            cur.close()
            data = build_xlsx(products, brands, categories)
            return resp(200, {'file': base64.b64encode(data).decode('ascii')})

        if action == 'wiring-xlsx':
            # scope: hot — только машины с большим разбросом цен (по умолчанию),
            # kit — где собирается комплект, all — вообще все
            scope = (params.get('scope') or 'hot').lower()
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(f"SELECT * FROM {schema()}.products ORDER BY sort_order, id")
            products = [row_to_product(r) for r in cur.fetchall()]
            cur.execute(f"SELECT name, models FROM {schema()}.brands ORDER BY sort_order, id")
            brands = [{'name': b['name'], 'models': b['models']} for b in cur.fetchall()]
            cur.execute(f"SELECT * FROM {schema()}.vehicle_wiring ORDER BY brand, model")
            saved = [
                {
                    'brand': s['brand'],
                    'model': s['model'],
                    'yearFrom': s['year_from'],
                    'yearTo': s['year_to'],
                    'mode': s['mode'],
                    'wireSlug': s['wire_slug'],
                    'reason': s['reason'],
                    'ask': s['ask'] or {},
                    # Кузов и руль тоже выгружаем: без них файл возвращался
                    # с пустыми колонками и загрузка стирала разметку
                    'wheel': s['wheel'] or '',
                    'bodies': s['bodies'] or [],
                }
                for s in cur.fetchall()
            ]
            cur.close()
            data = build_wiring_xlsx(products, brands, saved, scope)
            return resp(200, {'file': base64.b64encode(data).decode('ascii')})

        if action == 'wiring-import' and method == 'POST':
            raw = base64.b64decode(body.get('file') or '')
            parsed = parse_wiring_xlsx(raw)
            cur = conn.cursor()
            wires = updated = 0

            for w in parsed['wires']:
                if not (
                    w['tech']
                    or w['keeps']
                    or w['level']
                    or w['note']
                    or w['bodies']
                    or w['wheel']
                ):
                    continue
                cur.execute(
                    f"UPDATE {schema()}.products SET "
                    f"wire_tech = {qjson(w['tech'])}, "
                    f"wire_keeps = {qjson(w['keeps'])}, "
                    f"wire_level = {q(w['level'])}, "
                    f"wire_note = {q(w['note'])}, "
                    f"wire_bodies = {qjson(w['bodies'])}, "
                    f"wire_wheel = {q(w['wheel'])} "
                    f"WHERE slug = {q(w['slug'])}"
                )
                wires += cur.rowcount

            for c in parsed['cars']:
                # Строка уже есть на эти годы — обновляем, иначе добавляем.
                # Ключа в БД больше нет: у модели бывает несколько поколений
                cur.execute(
                    f"SELECT id FROM {schema()}.vehicle_wiring "
                    f"WHERE brand = {q(c['brand'])} AND model = {q(c['model'])} "
                    f"AND year_from = {qint(c['yearFrom'], 1990)} "
                    f"AND year_to = {qint(c['yearTo'], 2100)} LIMIT 1"
                )
                found = cur.fetchone()
                sets = (
                    f"mode = {q(c['mode'])}, "
                    f"wire_slug = {q(c['wireSlug'])}, "
                    f"reason = {q(c['reason'])}, "
                    f"ask = {qjson(c['ask'])}, "
                    f"bodies = {qjson(c['bodies'])}, "
                    f"wheel = {q(c['wheel'])}, updated_at = NOW()"
                )
                if found:
                    cur.execute(
                        f"UPDATE {schema()}.vehicle_wiring SET {sets} "
                        f"WHERE id = {found[0]}"
                    )
                else:
                    cur.execute(
                        f"INSERT INTO {schema()}.vehicle_wiring "
                        f"(brand, model, year_from, year_to, mode, wire_slug, "
                        f"reason, ask, bodies, wheel) "
                        f"VALUES ({q(c['brand'])}, {q(c['model'])}, "
                        f"{qint(c['yearFrom'], 1990)}, {qint(c['yearTo'], 2100)}, "
                        f"{q(c['mode'])}, {q(c['wireSlug'])}, {q(c['reason'])}, "
                        f"{qjson(c['ask'])}, {qjson(c['bodies'])}, "
                        f"{q(c['wheel'])})"
                    )
                updated += 1

            conn.commit()
            cur.close()
            return resp(
                200,
                {
                    'wires': wires,
                    'cars': updated,
                    'problems': parsed['problems'][:100],
                },
            )

        if action == 'supplier-scan':
            # Читаем прайс поставщика и сверяем с каталогом по артикулу:
            # что новинка, а что у нас уже есть
            raw = str(body.get('file', ''))
            if ',' in raw and raw.startswith('data:'):
                raw = raw.split(',', 1)[1]
            try:
                blob = base64.b64decode(raw)
            except Exception:
                return resp(400, {'error': 'Файл не читается'})
            try:
                items, idx = supplier.read_price(blob)
            except Exception:
                return resp(400, {'error': 'Это не таблица Excel'})
            if not items:
                return resp(400, {'error': 'В прайсе не нашлось строк с товарами'})
            if 'url' not in idx:
                return resp(400, {
                    'error': 'В прайсе нет колонки со ссылкой на товар — '
                             'без неё нечего открывать'
                })

            cur = conn.cursor()
            cur.execute(f"SELECT sku, slug FROM {schema()}.products WHERE sku <> ''")
            mine = {
                str(r[0]).strip().upper().replace(' ', ''): r[1]
                for r in cur.fetchall()
            }
            cur.close()

            new_count = 0
            for it in items:
                key = it['sku'].upper().replace(' ', '')
                it['slug'] = mine.get(key, '')
                it['exists'] = bool(it['slug'])
                if not it['exists']:
                    new_count += 1

            return resp(200, {
                'ok': True,
                'total': len(items),
                'new': new_count,
                'exists': len(items) - new_count,
                'noUrl': sum(1 for it in items if not it['url']),
                'items': items,
            })

        if action == 'supplier-collect':
            # Карточки обходим порциями: за один вызов функция успевает
            # немного, а прайс бывает и на тысячу строк
            items = body.get('items') or []
            if not items:
                return resp(400, {'error': 'Нечего собирать'})
            # Функции отведено около двух секунд: больше пяти карточек
            # за раз она не успевает и обрывается по таймауту
            items = items[:5]
            cur = conn.cursor()
            cur.execute(f"SELECT name, models FROM {schema()}.brands")
            brands = {r[0]: (r[1] or []) for r in cur.fetchall()}
            cur.close()
            rows = supplier.collect(items, brands)
            return resp(200, {
                'ok': True,
                'rows': rows,
                'failed': sum(1 for r in rows if not r.get('ok')),
            })

        if action == 'supplier-xlsx':
            rows = body.get('rows') or []
            if not rows:
                return resp(400, {'error': 'Нет данных для таблицы'})
            category = str(body.get('category') or '')
            data = supplier.build_xlsx(rows, category, DEFAULT_STOCK_NOTE)
            return resp(200, {'file': base64.b64encode(data).decode('ascii')})

        if action == 'import-xlsx':
            raw = str(body.get('file', ''))
            if ',' in raw and raw.startswith('data:'):
                raw = raw.split(',', 1)[1]
            try:
                blob = base64.b64decode(raw)
            except Exception:
                return resp(400, {'error': 'Файл не читается'})
            try:
                in_products, in_brands, in_categories = parse_xlsx(blob)
            except Exception:
                return resp(400, {'error': 'Это не таблица Excel или структура изменена'})
            if not in_products and not in_brands and not in_categories:
                return resp(400, {'error': 'В таблице не нашлось строк с товарами'})

            # Большой каталог грузим порциями: за один вызов функция успевает
            # обработать лишь часть строк, а на 1400 товарах упиралась в
            # таймаут — админка показывала ошибку, хотя часть уже сохранилась.
            # Браузер шлёт файл несколько раз, каждый раз со своим отрезком.
            total = len(in_products)
            try:
                offset = max(0, int(body.get('offset') or 0))
            except (TypeError, ValueError):
                offset = 0
            try:
                limit = int(body.get('limit') or 0)
            except (TypeError, ValueError):
                limit = 0
            chunk = in_products[offset:offset + limit] if limit > 0 else in_products
            done = offset + len(chunk) >= total

            # Марки и категории трогаем один раз — на первом отрезке
            stats = import_rows(
                conn,
                chunk,
                in_brands if offset == 0 else [],
                str(body.get('mode', 'merge')),
            )
            saved_cats = import_categories(conn, in_categories) if offset == 0 else 0

            if limit > 0:
                external = sum(
                    len([u for u in (p.get('images') or []) if _is_external_image(u)])
                    for p in chunk
                )
                return resp(
                    200,
                    {
                        'ok': True,
                        'total': total,
                        'processed': offset + len(chunk),
                        'done': done,
                        'brands': len(in_brands) if offset == 0 else 0,
                        'categories': saved_cats,
                        'external_images': external,
                        **stats,
                    },
                )
            # Сколько фото указано ссылками на чужие сайты — их можно перенести к нам
            external = sum(
                len([u for u in (p.get('images') or []) if _is_external_image(u)])
                for p in in_products
            )
            return resp(
                200,
                {
                    'ok': True,
                    'brands': len(in_brands),
                    'categories': saved_cats,
                    'external_images': external,
                    **stats,
                },
            )

        if action == 'orders':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            if method == 'GET':
                cur.execute(
                    f"SELECT * FROM {schema()}.orders ORDER BY created_at DESC LIMIT 500"
                )
                orders = [row_to_order(r) for r in cur.fetchall()]
                cur.close()
                return resp(200, {'orders': orders})
            if method == 'PUT':
                oid = qint(body.get('id'))
                if oid == 'NULL':
                    return resp(400, {'error': 'Не указана заявка'})
                status = str(body.get('status', 'new'))[:24]
                note = str(body.get('adminNote', ''))[:2000]
                cur.execute(
                    f"UPDATE {schema()}.orders SET status = {q(status)}, "
                    f"admin_note = {q(note)}, updated_at = NOW() WHERE id = {oid} RETURNING *"
                )
                row = cur.fetchone()
                conn.commit()
                cur.close()
                if not row:
                    return resp(404, {'error': 'Заявка не найдена'})
                return resp(200, {'order': row_to_order(row)})
            if method == 'DELETE':
                cur.execute(f"DELETE FROM {schema()}.orders WHERE id = {qint(params.get('id'))}")
                conn.commit()
                cur.close()
                return resp(200, {'ok': True})
            cur.close()
            return resp(400, {'error': 'Неизвестное действие'})

        if action == 'guides':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            if method == 'GET':
                cur.execute(f"SELECT * FROM {schema()}.guides ORDER BY sort_order, id")
                guides = [row_to_guide(r) for r in cur.fetchall()]
                cur.execute(
                    f"SELECT g.id AS gid, p.id AS pid FROM {schema()}.product_guides pg "
                    f"JOIN {schema()}.guides g ON g.id = pg.guide_id "
                    f"JOIN {schema()}.products p ON p.id = pg.product_id"
                )
                links: dict = {}
                for row in cur.fetchall():
                    links.setdefault(row['gid'], []).append(row['pid'])
                for g in guides:
                    g['productIds'] = links.get(g['id'], [])
                cur.close()
                return resp(200, {'guides': guides})

            if method == 'DELETE':
                gid = qint(params.get('id'))
                cur.execute(f"DELETE FROM {schema()}.product_guides WHERE guide_id = {gid}")
                cur.execute(f"DELETE FROM {schema()}.guides WHERE id = {gid}")
                conn.commit()
                cur.close()
                return resp(200, {'ok': True})

            if method in ('POST', 'PUT'):
                title = str(body.get('title', '')).strip()
                if not title:
                    return resp(400, {'error': 'Укажите заголовок'})
                slug = str(body.get('slug', '')).strip() or slugify(title)
                fields = {
                    'slug': q(slug),
                    'title': q(title),
                    'excerpt': q(str(body.get('excerpt', ''))),
                    'cover': q(str(body.get('cover', ''))),
                    'duration': q(str(body.get('duration', ''))),
                    'difficulty': q(str(body.get('difficulty', ''))),
                    'tools': qjson(body.get('tools') or []),
                    'blocks': qjson(body.get('blocks') or []),
                    'sort_order': qint(body.get('sortOrder'), 100),
                    'is_active': 'TRUE' if body.get('isActive', True) else 'FALSE',
                }
                if method == 'POST':
                    cur.execute(
                        f"INSERT INTO {schema()}.guides ({', '.join(fields.keys())}) "
                        f"VALUES ({', '.join(fields.values())}) RETURNING *"
                    )
                else:
                    gid = qint(body.get('id'))
                    if gid == 'NULL':
                        return resp(400, {'error': 'Не указана инструкция'})
                    sets = ', '.join(f"{k} = {v}" for k, v in fields.items())
                    cur.execute(
                        f"UPDATE {schema()}.guides SET {sets}, updated_at = NOW() "
                        f"WHERE id = {gid} RETURNING *"
                    )
                row = cur.fetchone()
                if not row:
                    conn.rollback()
                    cur.close()
                    return resp(404, {'error': 'Инструкция не найдена'})
                guide_id = row['id']
                cur.execute(
                    f"DELETE FROM {schema()}.product_guides WHERE guide_id = {guide_id}"
                )
                for pid in body.get('productIds') or []:
                    pid_sql = qint(pid)
                    if pid_sql != 'NULL':
                        cur.execute(
                            f"INSERT INTO {schema()}.product_guides (product_id, guide_id) "
                            f"VALUES ({pid_sql}, {guide_id}) ON CONFLICT DO NOTHING"
                        )
                conn.commit()
                out = row_to_guide(row)
                out['productIds'] = [int(p) for p in (body.get('productIds') or [])]
                cur.close()
                return resp(200, {'guide': out})
            cur.close()
            return resp(400, {'error': 'Неизвестное действие'})

        if method == 'GET':
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(
                f"SELECT * FROM {schema()}.products ORDER BY sort_order, id"
            )
            products = [row_to_product(r) for r in cur.fetchall()]
            cur.execute(
                f"SELECT name, models, model_bodies FROM {schema()}.brands ORDER BY sort_order, id"
            )
            brands = [
                {
                    'name': b['name'],
                    'models': b['models'],
                    'modelBodies': b['model_bodies'] or {},
                }
                for b in cur.fetchall()
            ]
            cur.execute(
                f"SELECT COUNT(*) AS c FROM {schema()}.orders WHERE status = 'new'"
            )
            new_orders = cur.fetchone()['c']
            cur.execute(
                f"SELECT COUNT(*) AS c FROM {schema()}.missing_fit_requests"
            )
            missing_fits = cur.fetchone()['c']
            cur.close()
            return resp(200, {
                'products': products,
                'brands': brands,
                'newOrders': new_orders,
                'missingFits': missing_fits,
            })

        if method == 'DELETE':
            pid = params.get('id', '')
            cur = conn.cursor()
            cur.execute(f"DELETE FROM {schema()}.product_guides WHERE product_id = {qint(pid)}")
            cur.execute(f"DELETE FROM {schema()}.products WHERE id = {qint(pid)}")
            conn.commit()
            cur.close()
            return resp(200, {'ok': True})

        if method in ('POST', 'PUT'):
            name = str(body.get('name', '')).strip()
            if not name:
                return resp(400, {'error': 'Укажите название'})
            slug = str(body.get('slug', '')).strip()
            if not slug:
                cur_s = conn.cursor()
                cur_s.execute(f"SELECT slug FROM {schema()}.products")
                taken = {r[0] for r in cur_s.fetchall()}
                cur_s.close()
                slug = make_slug(name, taken)
            category = str(body.get('category', '')).strip() or 'Другое'
            fields = {
                'slug': q(slug),
                'sku': q(str(body.get('sku', '')).strip() or slug.upper()),
                'name': q(name),
                'category': q(category),
                'price': qint(body.get('price'), 0),
                'old_price': qint(body.get('oldPrice')),
                'pro_price': qint(body.get('proPrice')),
                'ozon_url': q(str(body.get('ozonUrl', ''))),
                'wb_url': q(str(body.get('wbUrl', ''))),
                'install': q(str(body.get('install', ''))),
                'warranty': q(str(body.get('warranty', ''))),
                'year_from': qint(body.get('yearFrom'), 2010),
                'year_to': qint(body.get('yearTo'), 2026),
                'badge': q(body.get('badge') or None),
                'images': qjson(body.get('images') or []),
                'video_url': q(str(body.get('videoUrl') or '')[:500]),
                'description': qjson(body.get('description') or []),
                'specs': qjson(body.get('specs') or []),
                'kit': qjson(body.get('kit') or []),
                'notes': qjson(body.get('notes') or []),
                'extra': qjson(body.get('extra') or []),
                'extra_title': q(str(body.get('extraTitle') or '')[:160]),
                'fits': qjson(body.get('fits') or {}),
                # Пусто — значит берём умолчание категории
                'fit_mode': q(
                    body.get('fitMode') if body.get('fitMode') in FIT_MODES else ''
                ),
                'wire_tech': qjson(clean_wire_tech(body.get('wireTech') or {})),
                'wire_keeps': qjson(clean_wire_keeps(body.get('wireKeeps') or {})),
                'wire_level': q(
                    body.get('wireLevel')
                    if body.get('wireLevel') in WIRE_LEVELS
                    else ''
                ),
                'wire_note': q(str(body.get('wireNote') or '')[:600]),
                'wire_bodies': qjson(
                    [b for b in (body.get('wireBodies') or []) if b in BODY_TYPES]
                ),
                'wire_wheel': q(
                    body.get('wireWheel')
                    if body.get('wireWheel') in WHEEL_SIDES
                    else ''
                ),
                'frame_wires': qjson(
                    [str(x) for x in (body.get('frameWires') or []) if x][:20]
                ),
                'wire_included': 'TRUE' if body.get('wireIncluded') else 'FALSE',
                'sort_order': qint(body.get('sortOrder'), 100),
                'popularity': qint(body.get('popularity'), 0),
        'stock_qty': qint(body.get('stock'), 0),
        'stock_note': q(str(body.get('stockNote') or DEFAULT_STOCK_NOTE).strip() or DEFAULT_STOCK_NOTE),
                'is_active': 'TRUE' if body.get('isActive', True) else 'FALSE',
            }
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            if method == 'POST':
                cols = ', '.join(fields.keys())
                vals = ', '.join(fields.values())
                cur.execute(
                    f"INSERT INTO {schema()}.products ({cols}) VALUES ({vals}) RETURNING *"
                )
            else:
                pid = qint(body.get('id'))
                if pid == 'NULL':
                    return resp(400, {'error': 'Не указан товар'})
                sets = ', '.join(f"{k} = {v}" for k, v in fields.items())
                cur.execute(
                    f"UPDATE {schema()}.products SET {sets}, updated_at = NOW() "
                    f"WHERE id = {pid} RETURNING *"
                )
            row = cur.fetchone()
            conn.commit()
            cur.close()
            if not row:
                return resp(404, {'error': 'Товар не найден'})
            return resp(200, {'product': row_to_product(row)})

        return resp(400, {'error': 'Неизвестное действие'})
    finally:
        conn.close()