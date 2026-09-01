"""
Разбор прайса поставщика.

Поставщик присылает Excel: артикул, название, цена и ссылка на карточку.
Карточка знает больше — фотографии, описание, с какими машинами товар
совместим. Здесь мы обходим карточки и собираем из них таблицу того же
формата, что и выгрузка каталога: её остаётся проверить глазами и
загрузить обычным импортом.

Всё, чего нет в прайсе, берём со страницы; чего нет и там — оставляем
пустым, но помечаем цветом, чтобы человек увидел и решил сам.
"""

import io
import json
import re
import urllib.request
from concurrent.futures import ThreadPoolExecutor

UA = (
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
    '(KHTML, like Gecko) Chrome/120.0 Safari/537.36'
)

# Как поставщики называют колонки. Смотрим по вхождению, регистр не важен
COL_HINTS = {
    'sku': ('артикул', 'sku', 'код товара', 'код'),
    'name': ('название', 'наименование', 'товар', 'name'),
    'price': ('цена', 'стоимость', 'price', 'розница'),
    'url': ('ссылка', 'url', 'адрес', 'страница'),
}

# Названия марок, под которыми они встречаются в текстах поставщиков
ALIAS = {
    'vw': 'Volkswagen',
    'фольксваген': 'Volkswagen',
    'мерседес': 'Mercedes-Benz',
    'mercedes': 'Mercedes-Benz',
    'benz': 'Mercedes-Benz',
    'шкода': 'Skoda',
    'ssang': 'SsangYong',
    'ssangyong': 'SsangYong',
    'landrover': 'Land Rover',
    'ленд': 'Land Rover',
    'range': 'Land Rover',
    'лада': 'Lada (ВАЗ)',
    'lada': 'Lada (ВАЗ)',
    'ваз': 'Lada (ВАЗ)',
    'chevy': 'Chevrolet',
    'great': 'Great Wall',
    'greatwall': 'Great Wall',
    'тойота': 'Toyota',
    'ниссан': 'Nissan',
    'хендай': 'Hyundai',
    'хёндай': 'Hyundai',
    'киа': 'Kia',
}


def fetch(url: str, tries: int = 2) -> str:
    for attempt in range(tries):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': UA})
            with urllib.request.urlopen(req, timeout=12) as r:
                return r.read().decode('utf-8', 'ignore')
        except Exception:
            if attempt == tries - 1:
                return ''
    return ''


def read_price(blob: bytes) -> tuple:
    """Читаем прайс: ищем колонки по заголовку, а не по номеру —
    у каждого поставщика свой порядок."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}

    head = [str(c or '').strip().lower() for c in rows[0]]
    idx = {}
    for key, hints in COL_HINTS.items():
        for i, title in enumerate(head):
            if not title:
                continue
            if any(h in title for h in hints):
                idx.setdefault(key, i)
                break

    # Ссылку могли не подписать — ищем колонку, где лежат адреса
    if 'url' not in idx:
        for i in range(len(head)):
            sample = [r[i] for r in rows[1:12] if i < len(r) and r[i]]
            if sample and all(str(s).startswith('http') for s in sample):
                idx['url'] = i
                break

    items = []
    for r in rows[1:]:
        def cell(key):
            i = idx.get(key)
            return r[i] if i is not None and i < len(r) and r[i] is not None else ''

        sku = ' '.join(str(cell('sku')).split())
        url = str(cell('url')).strip()
        if not sku and not url:
            continue
        price = 0
        raw = str(cell('price')).replace(' ', '').replace(',', '.')
        m = re.search(r'\d+(?:\.\d+)?', raw)
        if m:
            price = int(float(m.group()))
        items.append({
            'sku': sku,
            'name': ' '.join(str(cell('name')).split()),
            'price': price,
            'url': url if url.startswith('http') else '',
        })
    return items, idx


def clean_html(raw: str) -> str:
    if not raw:
        return ''
    raw = re.sub(r'(?is)<(script|style).*?</\1>', ' ', raw)
    raw = re.sub(r'(?i)<br\s*/?>|</p>|</div>|</li>|</tr>', '\n', raw)
    txt = re.sub(r'<[^>]+>', ' ', raw)
    for a, b in (
        ('&nbsp;', ' '), ('&amp;', '&'), ('&quot;', '"'), ('&lt;', '<'),
        ('&gt;', '>'), ('&laquo;', '«'), ('&raquo;', '»'), ('&mdash;', '—'),
        ('&ndash;', '–'), ('&#39;', "'"),
    ):
        txt = txt.replace(a, b)
    lines = [re.sub(r'[ \t]+', ' ', x).strip() for x in txt.split('\n')]
    return '\n'.join(x for x in lines if x)


def js_block(html: str, var: str) -> str:
    m = re.search(re.escape(var) + r' = (\{.*?\});', html, re.S)
    if not m:
        return ''
    try:
        data = json.loads(m.group(1))
    except Exception:
        return ''
    return ' '.join(str(v) for v in data.values())


MUSOR = re.compile(
    r'^(купить на маркетплейсе|купить товар на|отзывы$|вопросы|доставка|оплата)',
    re.I,
)
SECTIONS = (
    'Описание', 'Совместимость', 'Совместимость по рамкам:',
    'Дополнительно', 'Комплектация', 'Характеристики',
)


def sections(html: str) -> dict:
    """Вкладка товара — сплошной текст: описание, совместимость и
    примечания идут подряд. Разбираем по заголовкам разделов."""
    txt = clean_html(js_block(html, 'var textAttributes'))
    if not txt:
        return {}
    out, cur = {}, None
    for line in txt.split('\n'):
        s = line.strip()
        if s in SECTIONS:
            cur = s
            continue
        if not s or MUSOR.match(s):
            continue
        out.setdefault(cur or 'Описание', []).append(s)
    return {k: '\n'.join(v) for k, v in out.items() if v}


def photos(html: str) -> list:
    urls = re.findall(
        r'https?://[^"\'\s>]+/uploads/images/items/max/[^"\'\s>]+?'
        r'\.(?:jpg|jpeg|png|webp)',
        html, re.I,
    )
    if not urls:
        # Запасной вариант — любые крупные картинки товара
        urls = re.findall(
            r'https?://[^"\'\s>]+/(?:upload|uploads|images|photo)[^"\'\s>]+?'
            r'\.(?:jpg|jpeg|png|webp)',
            html, re.I,
        )
    seen, out = set(), []
    for u in urls:
        if u not in seen and '/preview/' not in u and '/middle/' not in u:
            seen.add(u)
            out.append(u)
    return out[:15]


def specs(html: str) -> str:
    txt = clean_html(js_block(html, 'var mainAttributes'))
    rows = []
    for line in txt.split('\n'):
        if ':' in line:
            k, _, v = line.partition(':')
            k, v = k.strip(), v.strip()
            if k and v and len(k) < 60 and len(v) < 120:
                rows.append(f'{k}={v}')
    return '; '.join(rows)


YEARS = re.compile(r'\b(19[89]\d|20[0-4]\d)\b')
SHORT = re.compile(r'(?<![\d\w])([0-2]\d)\s*[-—–]\s*([0-2]\d)(?![\d\w])')


def years(text: str) -> tuple:
    ys = [int(y) for y in YEARS.findall(text)]
    # У поставщика годы часто сокращены: «A3 08 - 13» это 2008-2013
    for a, b in SHORT.findall(text):
        ys += [2000 + int(a), 2000 + int(b)]
    ys = [y for y in ys if 1985 <= y <= 2035]
    if not ys:
        return '', ''
    lo, hi = min(ys), max(ys)
    if '+' in text and hi < 2026:
        hi = 2026
    return lo, hi


def fits(text: str, brands: dict) -> str:
    """Ищем в тексте марки и модели из нашего справочника —
    чтобы товар попал в подбор по автомобилю."""
    if not text:
        return ''
    low = ' ' + re.sub(r'[^\w\s\-\+/]', ' ', text.lower()) + ' '
    found = {}
    for brand, models in brands.items():
        keys = {brand.lower()}
        for a, b in ALIAS.items():
            if b == brand:
                keys.add(a)
        if not any(
            re.search(r'(?<![\w])' + re.escape(k) + r'(?![\w])', low) for k in keys
        ):
            continue
        got = []
        for m in models:
            ml = str(m).lower()
            if ml and re.search(r'(?<![\w])' + re.escape(ml) + r'(?![\w])', low):
                got.append(m)
        # Убираем короткие названия, попавшие внутрь длинных: Tiggo при Tiggo 7
        got = [m for m in got
               if not any(m != o and str(m).lower() in str(o).lower() for o in got)]
        found[brand] = got
    return ' | '.join(
        f"{b}: {', '.join(ms)}" if ms else f'{b}: ' for b, ms in found.items()
    )


def collect_one(item: dict, brands: dict) -> dict:
    html = fetch(item['url']) if item.get('url') else ''
    out = dict(item)
    out.update({
        'images': [], 'description': '', 'specs': '',
        'fits': '', 'fitsSrc': '', 'yearFrom': '', 'yearTo': '',
        'ok': bool(html),
    })
    base = item.get('name', '')
    if html:
        m = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.S)
        if m:
            base = clean_html(m.group(1))[:200] or base
        d = sections(html)
        parts = [d[k] for k in ('Описание', 'Дополнительно', 'Комплектация')
                 if d.get(k)]
        fit_text = ' '.join(
            x for x in (d.get('Совместимость', ''),
                        d.get('Совместимость по рамкам:', '')) if x
        )
        out['images'] = photos(html)
        out['description'] = ' | '.join(' '.join(p.split('\n')) for p in parts)
        out['specs'] = specs(html)
        out['fitsSrc'] = ' '.join(fit_text.split('\n'))[:300]
        # Блок на странице точнее названия, но бывает пустым
        out['fits'] = fits(fit_text, brands) or fits(base, brands)
        out['yearFrom'], out['yearTo'] = years(fit_text or base)
    else:
        out['fits'] = fits(base, brands)
        out['yearFrom'], out['yearTo'] = years(base)
    out['name'] = ' '.join(base.split())
    return out


def collect(items: list, brands: dict, workers: int = 10) -> list:
    """Карточки грузим пачкой: по одной 261 товар занял бы минуты."""
    with ThreadPoolExecutor(max_workers=workers) as ex:
        return list(ex.map(lambda it: collect_one(it, brands), items))


XLS_COLS = [
    ('slug', 'Код (не менять)', 14),
    ('sku', 'Артикул', 16),
    ('name', 'Название', 52),
    ('category', 'Категория', 32),
    ('price', 'Цена', 11),
    ('oldPrice', 'Старая цена', 12),
    ('proPrice', 'Цена дилера', 12),
    ('warranty', 'Гарантия', 11),
    ('install', 'Установка', 14),
    ('ozonUrl', 'Ссылка Ozon', 14),
    ('wbUrl', 'Ссылка Wildberries', 14),
    ('yearFrom', 'Год с', 8),
    ('yearTo', 'Год по', 8),
    ('badge', 'Метка', 10),
    ('stock', 'Наличие (шт)', 12),
    ('stockNote', 'Если нет в наличии', 26),
    ('popularity', 'Популярность', 12),
    ('sortOrder', 'Порядок', 9),
    ('isActive', 'На сайте (да/нет)', 16),
    ('fits', 'Совместимость', 42),
    ('images', 'Фото (ссылки через ; — можно с других сайтов)', 46),
    ('description', 'Описание (абзацы через |)', 60),
    ('specs', 'Характеристики (Имя=Значение;)', 30),
    ('kit', 'Комплектация (через ;)', 22),
    ('_status', 'Есть у нас?', 13),
    ('_photos', 'Сколько фото', 12),
    ('_url', 'Страница поставщика', 34),
    ('_fitsSrc', 'Совместимость: текст с сайта', 42),
]

SERVICE = {'_status', '_photos', '_url', '_fitsSrc'}


def build_xlsx(rows: list, category: str, stock_note: str) -> bytes:
    """Собираем таблицу в формате выгрузки каталога, чтобы её можно
    было проверить и загрузить обычным импортом."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Товары'

    head = Font(bold=True, color='FFFFFF', size=10)
    fill_main = PatternFill('solid', fgColor='1B1B1B')
    fill_lock = PatternFill('solid', fgColor='E01B0C')
    fill_serv = PatternFill('solid', fgColor='6B7280')
    warn = PatternFill('solid', fgColor='FDE68A')
    have = PatternFill('solid', fgColor='D1FAE5')

    for i, (key, title, width) in enumerate(XLS_COLS, start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.font = head
        c.fill = (fill_lock if key == 'slug'
                  else fill_serv if key in SERVICE else fill_main)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = 'C2'

    for r, item in enumerate(rows, start=2):
        exists = item.get('exists')
        vals = {
            'slug': item.get('slug', ''),
            'sku': item.get('sku', ''),
            'name': item.get('name', ''),
            'category': item.get('category') or category,
            'price': item.get('price', 0),
            'oldPrice': None,
            'proPrice': None,
            'warranty': '',
            'install': '',
            'ozonUrl': '',
            'wbUrl': '',
            'yearFrom': item.get('yearFrom', ''),
            'yearTo': item.get('yearTo', ''),
            'badge': '',
            'stock': 0,
            'stockNote': stock_note,
            'popularity': 0,
            'sortOrder': 100,
            'isActive': 'да',
            'fits': item.get('fits', ''),
            'images': '; '.join(item.get('images') or []),
            'description': item.get('description', ''),
            'specs': item.get('specs', ''),
            'kit': '',
            '_status': 'уже есть' if exists else 'новинка',
            '_photos': len(item.get('images') or []),
            '_url': item.get('url', ''),
            '_fitsSrc': item.get('fitsSrc', ''),
        }
        for i, (key, _t, _w) in enumerate(XLS_COLS, start=1):
            cell = ws.cell(row=r, column=i, value=vals.get(key))
            cell.alignment = Alignment(
                vertical='top',
                wrap_text=key in ('name', 'description', 'fits', 'images'),
            )
            if key == '_status':
                cell.fill = have if exists else warn
            elif (key == '_photos' and not item.get('images')) \
                    or (key == 'fits' and not item.get('fits')) \
                    or (key == 'description' and not item.get('description')):
                cell.fill = warn

    ws.auto_filter.ref = (
        f'A1:{get_column_letter(len(XLS_COLS))}{len(rows) + 1}'
    )

    wsi = wb.create_sheet('Как пользоваться')
    wsi.column_dimensions['A'].width = 112
    info = [
        ('Что это за файл', True),
        ('Прайс поставщика, дополненный данными с карточек товаров: '
         'фото, описание, совместимость с машинами.', False),
        ('Колонки совпадают с выгрузкой каталога — файл можно загрузить '
         'через Каталог → Импорт.', False),
        ('', False),
        ('Колонка «Есть у нас?»', True),
        ('«новинка» — такого артикула в каталоге нет.', False),
        ('«уже есть» — товар у вас есть, у него подставлен код: '
         'при импорте обновится цена, а не создастся дубль.', False),
        ('', False),
        ('Что проверить перед загрузкой', True),
        ('1. Категорию — она у всех строк одна, поправьте где нужно.', False),
        ('2. Жёлтым помечено то, где чего-то не хватает: '
         'фото, описания или совместимости.', False),
        ('3. Наличие стоит 0 — товары «под заказ». '
         'Проставьте количество для того, что есть на складе.', False),
        ('4. Лишние строки просто удалите — загрузится только оставшееся.', False),
        ('', False),
        ('Фотографии', True),
        ('В колонке «Фото» — ссылки на сайт поставщика. При импорте сайт '
         'сам перенесёт их в наше хранилище.', False),
        ('', False),
        ('Серые колонки', True),
        ('Последние четыре — служебные, для проверки. '
         'Можно удалить, можно оставить: сайт их не читает.', False),
    ]
    for i, (text, bold) in enumerate(info, start=1):
        c = wsi.cell(row=i, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if bold:
            c.font = Font(bold=True, size=11)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
